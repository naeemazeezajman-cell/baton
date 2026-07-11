"""VAT Filing Engine — a strictly SEPARATE, REMOVABLE module (see REMOVING-VAT-ENGINE.md).

Everything VAT-specific lives in this one file: the vat_* tables, the router, Excel
template generation (openpyxl), upload parsing (pandas), reconciliation, the computation,
client emails, and sealing. Gated behind env VAT_ENGINE_ENABLED (default true) — when
false every endpoint 404s and the frontend hides all VAT UI.

The module's ONLY outward touch is duties.apply_completion() at the very end, completing
the linked duty through the existing machinery. Nothing outside this module imports it
except the two registration lines in app/main.py and alembic/env.py.
"""

import calendar
import io
import os
import re
import uuid
from datetime import date, datetime

from typing import Literal

from fastapi import APIRouter, Depends, File as FileParam, Form, HTTPException, UploadFile
from pydantic import BaseModel, EmailStr, Field
from sqlalchemy import (BigInteger, Boolean, CheckConstraint, Date, ForeignKey, Index, Integer,
                        Numeric, Text, UniqueConstraint, select, text)
from sqlalchemy.dialects.postgresql import JSONB, UUID
from sqlalchemy.orm import Mapped, Session, mapped_column
from sqlalchemy.types import TIMESTAMP

from .. import blobs, emails
from ..config import get_settings
from ..db import Base, get_db
from ..models import Client, Duty, File as FileModel, User
from ..security import current_user
from ..tenancy import get_scoped_or_404, tenant_select
from ..workflow import conflict, iso, now
from .duties import CADENCE_MONTHS, apply_completion, fmt_d
from .files import _download_token

TS = TIMESTAMP(timezone=True)
GEN_UUID = text("gen_random_uuid()")
NOW = text("now()")

STATUSES = ("ledgers_pending", "invoices_pending", "reconciled", "computation_draft",
            "awaiting_client_approval", "ready_to_file", "complete")
EMIRATES = ("Abu Dhabi", "Dubai", "Sharjah", "Ajman", "Umm Al Quwain", "Ras Al Khaimah", "Fujairah")
VAT_TOLERANCE = 0.01

LEDGER_COLUMNS = ["Invoice No", "Invoice Date", "Party Name", "TRN", "Emirate",
                  "Net Amount", "VAT Amount", "Type (Output/Input)", "Supply Category"]
REGISTER_COLUMNS = ["Invoice No", "Invoice Date", "Party", "Emirate", "Net", "VAT Amount", "Notes",
                    "Supply Category"]

SUPPLY_CATEGORIES = {"standard": "Standard (5%)", "zero_rated": "Zero-rated (0%)", "exempt": "Exempt",
                     "margin": "Margin scheme", "rcm_import": "RCM-Import",
                     "out_of_scope": "Out of scope (designated zone)"}

BUSINESS_CATEGORIES = ("Trading", "Services", "Real estate", "Used goods & vehicles", "Manufacturing",
                       "Logistics & transport", "Education", "Healthcare", "Financial services", "Other")
# the practitioner interview (wizard v2) — one stored answer per key, yes/no/not_sure + note
FLAG_KEYS = ("trn_confirmed", "has_zero_rated", "has_exempt", "designated_zone", "margin_scheme",
             "rcm_imports", "blocked_input_risk", "open_fta_matters")
FLAG_VALUE_LABELS = {"yes": "Yes", "no": "No", "not_sure": "Not sure"}

STAGGERS = {"jan_apr_jul_oct": "Jan/Apr/Jul/Oct", "feb_may_aug_nov": "Feb/May/Aug/Nov",
            "mar_jun_sep_dec": "Mar/Jun/Sep/Dec", "monthly": "Monthly"}
STAGGER_END_MONTHS = {"jan_apr_jul_oct": (1, 4, 7, 10), "feb_may_aug_nov": (2, 5, 8, 11),
                      "mar_jun_sep_dec": (3, 6, 9, 12)}

APPROVAL_BASES = {
    "evidence_upload": "written approval on file (upload)",
    "email_approval": "client approval received by email",
    "message_approval": "client approval received by message (WhatsApp/SMS)",
    "verbal_instruction": "verbal instruction to proceed",
}


# ---------- env gate ----------

def _enabled() -> bool:
    return os.getenv("VAT_ENGINE_ENABLED", "true").strip().lower() not in ("0", "false", "no", "off")


def require_enabled():
    if not _enabled():
        raise HTTPException(status_code=404, detail="Not Found")


router = APIRouter(prefix="/vat-engine", tags=["vat-engine"], dependencies=[Depends(require_enabled)])


# ---------- models (module-owned vat_* tables) ----------

class VatFiling(Base):
    __tablename__ = "vat_filings"
    __table_args__ = (CheckConstraint(f"status IN {STATUSES}", name="vat_filings_status_check"),
                      Index("ix_vat_filings_duty_id", "duty_id"))

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, server_default=GEN_UUID)
    tenant_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True))
    duty_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("duties.id"))
    client_id: Mapped[uuid.UUID | None] = mapped_column(ForeignKey("clients.id"))
    staff_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("users.id"))  # holder — always the staff
    period_start: Mapped[date] = mapped_column(Date)
    period_end: Mapped[date] = mapped_column(Date)
    prev_period_start: Mapped[date] = mapped_column(Date)  # window rule boundary
    status: Mapped[str] = mapped_column(Text, server_default="ledgers_pending")
    ledger_file: Mapped[dict | None] = mapped_column(JSONB)
    invoice_file: Mapped[dict | None] = mapped_column(JSONB)
    invoice_evidence: Mapped[list] = mapped_column(JSONB, server_default=text("'[]'"))
    recon: Mapped[dict | None] = mapped_column(JSONB)
    computation: Mapped[dict | None] = mapped_column(JSONB)
    client_approval: Mapped[dict | None] = mapped_column(JSONB)
    fta_ack: Mapped[dict | None] = mapped_column(JSONB)
    created_at: Mapped[datetime] = mapped_column(TS, server_default=NOW)
    completed_at: Mapped[datetime | None] = mapped_column(TS)


class VatFilingItem(Base):
    __tablename__ = "vat_filing_items"
    __table_args__ = (Index("ix_vat_filing_items_filing_id", "filing_id"),)

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, server_default=GEN_UUID)
    tenant_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True))
    filing_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("vat_filings.id"))
    source: Mapped[str] = mapped_column(Text)  # ledger | invoice
    row_no: Mapped[int] = mapped_column(BigInteger)  # excel row for error/trace messages
    invoice_no: Mapped[str] = mapped_column(Text)
    invoice_no_norm: Mapped[str] = mapped_column(Text)
    invoice_date: Mapped[date] = mapped_column(Date)
    party: Mapped[str] = mapped_column(Text)
    trn: Mapped[str | None] = mapped_column(Text)
    emirate: Mapped[str] = mapped_column(Text)
    net: Mapped[float] = mapped_column(Numeric(14, 2))
    vat: Mapped[float] = mapped_column(Numeric(14, 2))
    type_: Mapped[str | None] = mapped_column("type", Text)  # Output | Input (ledger only)
    category: Mapped[str] = mapped_column(Text, server_default="standard")  # supply category key
    origin: Mapped[str] = mapped_column(Text, server_default="register")  # register | ai_extracted
    notes: Mapped[str | None] = mapped_column(Text)
    bucket: Mapped[str | None] = mapped_column(Text)  # matched | ledger_only | invoice_only | out_of_window
    resolution: Mapped[dict | None] = mapped_column(JSONB)  # {action, reason, by, at}
    included: Mapped[bool] = mapped_column(Boolean, server_default=text("true"))


class VatFilingEvent(Base):
    __tablename__ = "vat_filing_events"

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)
    tenant_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True))
    filing_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("vat_filings.id"))
    at: Mapped[datetime] = mapped_column(TS, server_default=NOW)
    by_user: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))
    text_: Mapped[str] = mapped_column("text", Text)


class VatClientProfile(Base):
    """The engine's memory: one VAT profile per client — nature of business plus the
    compliance-relevant flags, each with an optional note. Every edit appends to the
    updated log and bumps the version; filings record which version applied."""

    __tablename__ = "vat_client_profiles"
    __table_args__ = (UniqueConstraint("tenant_id", "client_id"),)

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, server_default=GEN_UUID)
    tenant_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True))
    client_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("clients.id"))
    nature_of_business: Mapped[str | None] = mapped_column(Text)
    business_category: Mapped[str] = mapped_column(Text)
    tax_period_stagger: Mapped[str | None] = mapped_column(Text)  # STAGGERS key — drives period derivation
    flags: Mapped[dict] = mapped_column(JSONB, server_default=text("'{}'"))  # {key: {value: yes|no|not_sure, note}}
    other_notes: Mapped[str | None] = mapped_column(Text)
    version: Mapped[int] = mapped_column(Integer, server_default=text("1"))
    created_by: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))
    created_at: Mapped[datetime] = mapped_column(TS, server_default=NOW)
    updated: Mapped[list] = mapped_column(JSONB, server_default=text("'[]'"))  # [{at, by, by_name, changes}]


class VatExtractionDraft(Base):
    """AI-extracted invoice fields awaiting the MANDATORY human review. Only confirmed
    drafts become register items (origin=ai_extracted); unconfirmed drafts never reconcile."""

    __tablename__ = "vat_extraction_drafts"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, server_default=GEN_UUID)
    tenant_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True))
    filing_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("vat_filings.id"))
    file_id: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))  # source document (kept as evidence)
    file_name: Mapped[str] = mapped_column(Text)
    status: Mapped[str] = mapped_column(Text, server_default="extracted")  # extracted | failed | confirmed
    fields: Mapped[dict] = mapped_column(JSONB, server_default=text("'{}'"))  # {name: {value, confidence}}
    error: Mapped[str | None] = mapped_column(Text)
    created_by: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))
    created_at: Mapped[datetime] = mapped_column(TS, server_default=NOW)
    reviewed_by: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))
    reviewed_at: Mapped[datetime | None] = mapped_column(TS)


class VatClientRequest(Base):
    __tablename__ = "vat_client_requests"

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)
    tenant_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True))
    filing_id: Mapped[uuid.UUID] = mapped_column(ForeignKey("vat_filings.id"))
    kind: Mapped[str] = mapped_column(Text)  # ledger | invoices | missing_invoice | computation
    item_id: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))
    to_email: Mapped[str] = mapped_column(Text)
    subject: Mapped[str] = mapped_column(Text)
    sent_at: Mapped[datetime] = mapped_column(TS, server_default=NOW)
    by_user: Mapped[uuid.UUID | None] = mapped_column(UUID(as_uuid=True))


# ---------- schemas ----------

class OpenIn(BaseModel):
    duty_id: uuid.UUID


class MailIn(BaseModel):
    to: EmailStr
    subject: str = Field(min_length=1)
    body: str = Field(min_length=1)


class RequestFromClientIn(MailIn):
    kind: str  # ledger | invoices


class ReasonIn(BaseModel):
    reason: str = Field(min_length=1)


class FlagIn(BaseModel):
    value: Literal["yes", "no", "not_sure"]
    note: str | None = None


class ProfileIn(BaseModel):
    nature_of_business: str = ""
    business_category: str
    tax_period_stagger: Literal["jan_apr_jul_oct", "feb_may_aug_nov", "mar_jun_sep_dec", "monthly"] | None = None
    flags: dict[str, FlagIn] = {}
    other_notes: str | None = None


class ConfirmComputationIn(BaseModel):
    confirmations: list[str] = []
    warning_note: str = ""


class ExtractionRowIn(BaseModel):
    draft_id: uuid.UUID
    invoice_no: str = Field(min_length=1)
    invoice_date: str  # ISO date
    party: str = Field(min_length=1)
    emirate: str
    net: float
    vat: float
    currency: str = "AED"
    conversion_note: str = ""  # mandatory when currency != AED (manual conversion)


class ConfirmExtractionsIn(BaseModel):
    rows: list[ExtractionRowIn] = Field(min_length=1)


class AddToLedgerIn(BaseModel):
    invoice_no: str = Field(min_length=1)
    invoice_date: str  # ISO date
    party: str = Field(min_length=1)
    emirate: str
    net: float
    vat: float
    type: Literal["Output", "Input"]
    category: str = "standard"
    note: str = Field(min_length=1)  # MANDATORY correction note — the client's books get fixed too


# ---------- helpers ----------

def _log(db: Session, f: VatFiling, by_user, txt: str):
    db.add(VatFilingEvent(tenant_id=f.tenant_id, filing_id=f.id, by_user=by_user, text_=txt))


def _get(db: Session, fid: uuid.UUID, user: User) -> VatFiling:
    return get_scoped_or_404(db, VatFiling, fid, user)


def _require_staff_open(f: VatFiling, user: User):
    if f.status == "complete":
        raise conflict("This filing is complete — the trail is sealed")
    if f.staff_id != user.id:
        raise conflict("Only the responsible staff member can act on this filing")


def _require_status(f: VatFiling, *allowed: str):
    if f.status not in allowed:
        raise conflict(f"Filing is at stage '{f.status}' — this action requires {' / '.join(allowed)}")


def _period_label(f: VatFiling) -> str:
    return f"{f.period_start:%d %b %Y} – {f.period_end:%d %b %Y}"


def _month_shift(year: int, month: int, months_back: int) -> tuple[int, int]:
    m = year * 12 + (month - 1) - months_back
    return m // 12, m % 12 + 1


def derive_period(duty: Duty, profile=None) -> tuple[date, date, date]:
    """Filing period from the duty's schedule. When the client's VAT profile records a tax
    period stagger, the stagger drives the period: the period ends on the latest stagger
    month-end before the due month (monthly stagger → one-month periods). Without a
    profile, the period simply ends the month before the due month and spans one cadence.
    The window-rule boundary is one further period back."""
    months = CADENCE_MONTHS.get(duty.cadence, 3)
    stagger = getattr(profile, "tax_period_stagger", None) if profile else None
    if stagger == "monthly":
        months = 1
    due = duty.next_due
    end_y, end_m = _month_shift(due.year, due.month, 1)  # month before the due month
    if stagger in STAGGER_END_MONTHS:
        months = 3
        while end_m not in STAGGER_END_MONTHS[stagger]:
            end_y, end_m = _month_shift(end_y, end_m, 1)
    period_end = date(end_y, end_m, calendar.monthrange(end_y, end_m)[1])
    y1, m1 = _month_shift(end_y, end_m, months - 1)
    y2, m2 = _month_shift(end_y, end_m, 2 * months - 1)
    return date(y1, m1, 1), period_end, date(y2, m2, 1)


def _store_bytes(db: Session, user: User, entity: str, entity_id, name: str, data: bytes) -> FileModel:
    """Store generated bytes as a files row (registry files use entity='client')."""
    path = blobs.blob_path_for(user.tenant_id, entity, name)
    blobs.save_blob(path, data)
    row = FileModel(tenant_id=user.tenant_id, entity=entity, entity_id=entity_id,
                    name=name, size=len(data), blob_path=path, uploaded_by=user.id)
    db.add(row)
    db.flush()
    return row


def _store_upload(db: Session, user: User, entity: str, entity_id, upload: UploadFile) -> FileModel:
    return _store_bytes(db, user, entity, entity_id, upload.filename, upload.file.read())


def _file_link(f: FileModel) -> str:
    url = blobs.sas_link(f.blob_path)
    if url is None:
        url = f"{get_settings().FRONTEND_ORIGIN}/files/{f.id}/download?token={_download_token(f.id)}"
    return url


def serialize(db: Session, f: VatFiling, detail: bool = False) -> dict:
    duty = db.get(Duty, f.duty_id)
    client = db.get(Client, f.client_id) if f.client_id else None
    staff = db.get(User, f.staff_id)
    out = {
        "id": f.id, "duty_id": f.duty_id, "client_id": f.client_id,
        "client_name": client.name if client else (duty.client_name if duty else None),
        "client_ref": client.ref if client else None,
        "client_contact": client.contact if client else (duty.contact if duty else None),
        "staff_id": f.staff_id, "staff_name": staff.name if staff else None,
        "service": duty.service if duty else "VAT Filing",
        "period_start": f.period_start, "period_end": f.period_end,
        "prev_period_start": f.prev_period_start, "period_label": _period_label(f),
        "status": f.status, "holder": f.staff_id,
        "ledger_file": f.ledger_file, "invoice_file": f.invoice_file,
        "invoice_evidence": f.invoice_evidence, "recon": f.recon,
        "computation": f.computation, "client_approval": f.client_approval, "fta_ack": f.fta_ack,
        "created_at": f.created_at, "completed_at": f.completed_at,
        "duty_next_due": duty.next_due if duty else None,
    }
    if detail:
        items = db.scalars(select(VatFilingItem).where(VatFilingItem.filing_id == f.id)
                           .order_by(VatFilingItem.source, VatFilingItem.row_no)).all()
        out["items"] = [{
            "id": i.id, "source": i.source, "row_no": i.row_no, "invoice_no": i.invoice_no,
            "invoice_date": i.invoice_date, "party": i.party, "trn": i.trn, "emirate": i.emirate,
            "net": float(i.net), "vat": float(i.vat), "type": i.type_, "category": i.category,
            "origin": i.origin, "notes": i.notes,
            "bucket": i.bucket, "resolution": i.resolution, "included": i.included,
        } for i in items]
        events = db.scalars(select(VatFilingEvent).where(VatFilingEvent.filing_id == f.id)
                            .order_by(VatFilingEvent.at, VatFilingEvent.id)).all()
        out["events"] = [{"at": e.at, "by": e.by_user, "text": e.text_} for e in events]
        reqs = db.scalars(select(VatClientRequest).where(VatClientRequest.filing_id == f.id)
                          .order_by(VatClientRequest.sent_at, VatClientRequest.id)).all()
        out["client_requests"] = [{"id": r.id, "kind": r.kind, "item_id": r.item_id, "to": r.to_email,
                                   "subject": r.subject, "sent_at": r.sent_at, "by": r.by_user} for r in reqs]
        out["unresolved_differences"] = _unresolved_count(items)
        out["profile"] = _serialize_profile(_get_profile(db, f.tenant_id, f.client_id))
        drafts = db.scalars(select(VatExtractionDraft).where(VatExtractionDraft.filing_id == f.id)
                            .order_by(VatExtractionDraft.created_at, VatExtractionDraft.id)).all()
        out["extraction_drafts"] = [_serialize_draft(d) for d in drafts]
    return out


def _unresolved_count(items) -> int:
    return sum(1 for i in items
               if i.bucket in ("ledger_only", "invoice_only")
               and (i.resolution or {}).get("action") not in ("excluded", "resolved"))


# ---------- client VAT profile (the engine's memory) ----------

def _get_profile(db: Session, tenant_id, client_id) -> VatClientProfile | None:
    if client_id is None:
        return None
    return db.scalar(select(VatClientProfile).where(VatClientProfile.tenant_id == tenant_id,
                                                    VatClientProfile.client_id == client_id))


def _serialize_profile(p: VatClientProfile | None) -> dict | None:
    if p is None:
        return None
    return {"id": p.id, "client_id": p.client_id, "nature_of_business": p.nature_of_business,
            "business_category": p.business_category,
            "tax_period_stagger": p.tax_period_stagger,
            "tax_period_stagger_label": STAGGERS.get(p.tax_period_stagger),
            "flags": p.flags, "other_notes": p.other_notes,
            "version": p.version, "created_by": p.created_by, "created_at": p.created_at,
            "updated": p.updated}


def _require_profile_editor(db: Session, user: User, client_id):
    """Admin / Manager / staff assigned a VAT duty for this client."""
    if user.role in ("Admin", "Manager"):
        return
    has_duty = db.scalar(select(Duty.id).where(Duty.tenant_id == user.tenant_id,
                                               Duty.client_id == client_id,
                                               Duty.staff_id == user.id, Duty.kind == "vat"))
    if has_duty is None:
        raise conflict("Only Admin, Manager, or the assigned VAT staff can edit this client's VAT profile")


def _normalize_flags(flags: dict[str, FlagIn]) -> dict:
    out = {}
    for k in FLAG_KEYS:
        f = flags.get(k)
        out[k] = {"value": f.value, "note": (f.note or "").strip() or None} if f else {"value": "no", "note": None}
    unknown = set(flags) - set(FLAG_KEYS)
    if unknown:
        raise HTTPException(status_code=422, detail=f"Unknown profile flag(s): {sorted(unknown)}")
    return out


def _log_to_open_filings(db: Session, user: User, client_id, txt: str):
    for f in db.scalars(select(VatFiling).where(VatFiling.tenant_id == user.tenant_id,
                                                VatFiling.client_id == client_id,
                                                VatFiling.status != "complete")).all():
        _log(db, f, user.id, txt)


def _realign_open_periods(db: Session, user: User, client_id, profile: VatClientProfile):
    """The stagger drives every deadline: open filings that haven't started collecting yet
    re-derive their period when the profile's stagger is recorded or changed."""
    for f in db.scalars(select(VatFiling).where(VatFiling.tenant_id == user.tenant_id,
                                                VatFiling.client_id == client_id,
                                                VatFiling.status == "ledgers_pending")).all():
        duty = db.get(Duty, f.duty_id)
        if duty is None:
            continue
        ps, pe, pps = derive_period(duty, profile)
        if (ps, pe, pps) != (f.period_start, f.period_end, f.prev_period_start):
            f.period_start, f.period_end, f.prev_period_start = ps, pe, pps
            _log(db, f, user.id, f"Filing period re-aligned to the profile's tax period stagger "
                                 f"({STAGGERS.get(profile.tax_period_stagger, '—')}): {_period_label(f)}. "
                                 f"Window rule boundary now {pps:%d %b %Y}.")


@router.get("/clients/{client_id}/profile")
def get_profile(client_id: uuid.UUID, user: User = Depends(current_user), db: Session = Depends(get_db)):
    get_scoped_or_404(db, Client, client_id, user)
    p = _get_profile(db, user.tenant_id, client_id)
    if p is None:
        raise HTTPException(status_code=404, detail="No VAT profile recorded for this client yet")
    return _serialize_profile(p)


@router.post("/clients/{client_id}/profile", status_code=201)
def create_profile(client_id: uuid.UUID, body: ProfileIn, user: User = Depends(current_user),
                   db: Session = Depends(get_db)):
    client = get_scoped_or_404(db, Client, client_id, user)
    _require_profile_editor(db, user, client_id)
    if _get_profile(db, user.tenant_id, client_id):
        raise conflict("A VAT profile already exists for this client — edit it instead")
    if body.business_category not in BUSINESS_CATEGORIES:
        raise HTTPException(status_code=422, detail=f"business_category must be one of {BUSINESS_CATEGORIES}")
    flags_n = _normalize_flags(body.flags)
    p = VatClientProfile(tenant_id=user.tenant_id, client_id=client_id,
                         nature_of_business=body.nature_of_business.strip() or None,
                         business_category=body.business_category,
                         tax_period_stagger=body.tax_period_stagger,
                         flags=flags_n,
                         other_notes=(body.other_notes or "").strip() or None,
                         created_by=user.id)
    # v1 history entry — every answer, including "Not sure", is on record from day one
    v1_changes = [{"field": "business_category", "old": None, "new": body.business_category, "note": None},
                  {"field": "tax_period_stagger", "old": None,
                   "new": STAGGERS.get(body.tax_period_stagger), "note": None}]
    if body.nature_of_business.strip():
        v1_changes.append({"field": "nature_of_business", "old": None,
                           "new": body.nature_of_business.strip(), "note": None})
    for k in FLAG_KEYS:
        v1_changes.append({"field": k, "old": None, "new": FLAG_VALUE_LABELS[flags_n[k]["value"]],
                           "note": flags_n[k]["note"]})
    p.updated = [{"version": 1, "at": iso(now()), "by": str(user.id), "by_name": user.name,
                  "changes": v1_changes}]
    db.add(p)
    db.flush()
    active = [k for k in FLAG_KEYS if p.flags[k]["value"] in ("yes", "not_sure")]
    unsure = [k for k in FLAG_KEYS if p.flags[k]["value"] == "not_sure"]
    _log_to_open_filings(db, user, client_id,
                         f"VAT client profile v1 recorded by {user.name} for {client.name} "
                         f"({body.business_category}"
                         + (f", stagger {STAGGERS[body.tax_period_stagger]}" if body.tax_period_stagger else "")
                         + f") — flags: {', '.join(active) or 'none'}"
                         + (f"; NOT SURE (confirm with client): {', '.join(unsure)}" if unsure else "")
                         + ". Applied to this filing.")
    if body.tax_period_stagger:
        _realign_open_periods(db, user, client_id, p)
    db.commit()
    return _serialize_profile(p)


@router.patch("/clients/{client_id}/profile")
def update_profile(client_id: uuid.UUID, body: ProfileIn, user: User = Depends(current_user),
                   db: Session = Depends(get_db)):
    get_scoped_or_404(db, Client, client_id, user)
    _require_profile_editor(db, user, client_id)
    p = _get_profile(db, user.tenant_id, client_id)
    if p is None:
        raise HTTPException(status_code=404, detail="No VAT profile recorded for this client yet")
    if body.business_category not in BUSINESS_CATEGORIES:
        raise HTTPException(status_code=422, detail=f"business_category must be one of {BUSINESS_CATEGORIES}")
    new_flags = _normalize_flags(body.flags)
    # structured version rows: {field, old → new, note}
    changes = []
    if (p.nature_of_business or "") != (body.nature_of_business.strip() or ""):
        changes.append({"field": "nature_of_business", "old": p.nature_of_business or "—",
                        "new": body.nature_of_business.strip() or "—", "note": None})
    if p.business_category != body.business_category:
        changes.append({"field": "business_category", "old": p.business_category,
                        "new": body.business_category, "note": None})
    if p.tax_period_stagger != body.tax_period_stagger:
        changes.append({"field": "tax_period_stagger", "old": STAGGERS.get(p.tax_period_stagger, "—"),
                        "new": STAGGERS.get(body.tax_period_stagger, "—"), "note": None})
    for k in FLAG_KEYS:
        old, new = p.flags.get(k) or {"value": "no", "note": None}, new_flags[k]
        if old.get("value") != new["value"]:
            changes.append({"field": k, "old": FLAG_VALUE_LABELS[old.get("value", "no")],
                            "new": FLAG_VALUE_LABELS[new["value"]], "note": new["note"]})
        elif (old.get("note") or None) != new["note"]:
            changes.append({"field": k, "old": old.get("note") or "—", "new": new["note"] or "—",
                            "note": "note changed"})
    if (p.other_notes or None) != ((body.other_notes or "").strip() or None):
        changes.append({"field": "other_notes", "old": p.other_notes or "—",
                        "new": (body.other_notes or "").strip() or "—", "note": None})
    if not changes:
        return _serialize_profile(p)
    stagger_changed = p.tax_period_stagger != body.tax_period_stagger
    p.nature_of_business = body.nature_of_business.strip() or None
    p.business_category = body.business_category
    p.tax_period_stagger = body.tax_period_stagger
    p.flags = new_flags
    p.other_notes = (body.other_notes or "").strip() or None
    p.version = p.version + 1
    p.updated = [*p.updated, {"at": iso(now()), "by": str(user.id), "by_name": user.name,
                              "version": p.version, "changes": changes}]
    rendered = "; ".join(
        f"{c['field']} {c['old']} → {c['new']}" + (f' — note: "{c["note"]}"' if c["note"] else "")
        for c in changes)
    _log_to_open_filings(db, user, client_id,
                         f"Profile updated to v{p.version} by {user.name}: {rendered}")
    if stagger_changed:
        _realign_open_periods(db, user, client_id, p)
    db.commit()
    return _serialize_profile(p)


# ---------- status probe (frontend hides all UI on 404) ----------

@router.get("/status")
def status(user: User = Depends(current_user)):
    return {"enabled": True}


# ---------- Excel templates (openpyxl, generated on the fly) ----------

def _template_workbook(columns: list[str], instruction: str, validations: dict[str, list[str]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    c = ws.cell(row=1, column=1, value=instruction)
    c.font = Font(italic=True, size=9, color="666666")
    for idx, name in enumerate(columns, start=1):
        h = ws.cell(row=2, column=idx, value=name)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="14606B")
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(name) + 4)
    for col_name, options in validations.items():
        col_idx = columns.index(col_name) + 1
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True,
                            showErrorMessage=True, errorTitle="Invalid value",
                            error=f"Pick one of: {', '.join(options)}")
        ws.add_data_validation(dv)
        letter = get_column_letter(col_idx)
        dv.add(f"{letter}3:{letter}1000")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


LEDGER_INSTRUCTION = ("VAT Ledger — one row per invoice line. Keep the columns exactly as given "
                      "(the upload is validated against them). Dates as dd/mm/yyyy or Excel dates. "
                      "Type: Output = sales, Input = purchases. One file covers both two-ledger and "
                      "single-ledger bookkeeping — mark each row's Type. "
                      "Supply Category defaults to Standard (5%) when left blank.")
REGISTER_INSTRUCTION = ("Invoice Register — one row per issued invoice. Keep the columns exactly as "
                        "given (the upload is validated against them). Dates as dd/mm/yyyy or Excel dates. "
                        "Supply Category defaults to Standard (5%) when left blank.")


def _ledger_template_bytes() -> bytes:
    return _template_workbook(LEDGER_COLUMNS, LEDGER_INSTRUCTION,
                              {"Type (Output/Input)": ["Output", "Input"], "Emirate": list(EMIRATES),
                               "Supply Category": list(SUPPLY_CATEGORIES.values())})


def _register_template_bytes() -> bytes:
    return _template_workbook(REGISTER_COLUMNS, REGISTER_INSTRUCTION,
                              {"Emirate": list(EMIRATES),
                               "Supply Category": list(SUPPLY_CATEGORIES.values())})


def _xlsx_response(data: bytes, name: str):
    from fastapi.responses import Response
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{name}"'})


@router.get("/templates/ledger")
def ledger_template(user: User = Depends(current_user)):
    return _xlsx_response(_ledger_template_bytes(), "VAT Ledger Template.xlsx")


@router.get("/templates/invoice-register")
def register_template(user: User = Depends(current_user)):
    return _xlsx_response(_register_template_bytes(), "Invoice Register Template.xlsx")


# ---------- upload parsing (pandas; hard-fail with row-level errors) ----------

def _parse_category(v) -> str | None:
    """Supply Category cell → category key; blank defaults to standard; unknown → None (row error)."""
    if v is None or not str(v).strip():
        return "standard"
    s = str(v).strip().lower()
    for key, label in SUPPLY_CATEGORIES.items():
        if s in (label.lower(), key, key.replace("_", "-"), key.replace("_", " ")):
            return key
    for prefix, key in (("standard", "standard"), ("zero", "zero_rated"), ("exempt", "exempt"),
                        ("margin", "margin"), ("rcm", "rcm_import")):
        if s.startswith(prefix):
            return key
    return None


def _norm_invoice_no(v) -> str:
    s = str(v).strip()
    if s.endswith(".0"):  # Excel numeric cells round-trip as floats
        s = s[:-2]
    return re.sub(r"\s+", "", s).upper()


def _parse_upload(data: bytes, columns: list[str], is_ledger: bool) -> list[dict]:
    import pandas as pd

    try:
        df = pd.read_excel(io.BytesIO(data), header=1)
    except Exception:
        raise HTTPException(status_code=422, detail={
            "reason": "The file could not be read as an Excel workbook — download the template and fill it in.",
            "errors": []})
    got = [str(c).strip() for c in df.columns[:len(columns)]]
    if got != columns:
        raise HTTPException(status_code=422, detail={
            "reason": "Columns don't match the template — download the template and keep its columns exactly.",
            "errors": [f"Expected: {' | '.join(columns)}", f"Got: {' | '.join(got) or '(no header row found)'}"]})
    df = df[df.columns[:len(columns)]]
    df = df.dropna(how="all")
    rows, errors = [], []
    for idx, raw in df.iterrows():
        excel_row = idx + 3  # header on row 2, data starts row 3
        vals = dict(zip(columns, raw.tolist()))

        def val(col):
            v = vals.get(col)
            return None if pd.isna(v) else v

        inv_no = val("Invoice No")
        if inv_no is None or not str(inv_no).strip():
            errors.append(f"Row {excel_row}: Invoice No is required")
            continue
        try:
            d = pd.to_datetime(val("Invoice Date"), dayfirst=True)
            if pd.isna(d):
                raise ValueError
            inv_date = d.date()
        except Exception:
            errors.append(f"Row {excel_row}: Invoice Date {val('Invoice Date')!r} is not a valid date")
            continue
        party_col = "Party Name" if is_ledger else "Party"
        party = str(val(party_col) or "").strip()
        if not party:
            errors.append(f"Row {excel_row}: {party_col} is required")
            continue
        emirate = str(val("Emirate") or "").strip().title().replace("Al Quwain", "Al Quwain")
        matched_emirate = next((e for e in EMIRATES if e.lower() == emirate.lower()), None)
        if matched_emirate is None:
            errors.append(f"Row {excel_row}: Emirate {val('Emirate')!r} must be one of {', '.join(EMIRATES)}")
            continue
        net_col = "Net Amount" if is_ledger else "Net"
        try:
            net = round(float(val(net_col)), 2)
            vat = round(float(val("VAT Amount")), 2)
        except (TypeError, ValueError):
            errors.append(f"Row {excel_row}: {net_col} and VAT Amount must be numbers")
            continue
        row_type = None
        if is_ledger:
            row_type = str(val("Type (Output/Input)") or "").strip().capitalize()
            if row_type not in ("Output", "Input"):
                errors.append(f"Row {excel_row}: Type must be Output or Input, got {val('Type (Output/Input)')!r}")
                continue
        category = _parse_category(val("Supply Category"))
        if category is None:
            errors.append(f"Row {excel_row}: Supply Category {val('Supply Category')!r} must be one of "
                          f"{', '.join(SUPPLY_CATEGORIES.values())} (blank = Standard)")
            continue
        rows.append({
            "row_no": excel_row, "invoice_no": str(inv_no).strip(),
            "invoice_no_norm": _norm_invoice_no(inv_no), "invoice_date": inv_date,
            "party": party, "trn": str(val("TRN")).strip() if is_ledger and val("TRN") is not None else None,
            "emirate": matched_emirate, "net": net, "vat": vat, "type": row_type, "category": category,
            "notes": str(val("Notes")).strip() if not is_ledger and val("Notes") is not None else None,
        })
    if errors:
        raise HTTPException(status_code=422, detail={
            "reason": f"Template mismatch — {len(errors)} row error(s). Fix the file and re-upload; "
                      f"that's what the template is for.",
            "errors": errors})
    if not rows:
        raise HTTPException(status_code=422, detail={"reason": "The file has no data rows.", "errors": []})
    return rows


# ---------- open / read ----------

@router.get("/filings")
def list_filings(user: User = Depends(current_user), db: Session = Depends(get_db)):
    q = tenant_select(VatFiling, user)
    rows = db.scalars(q.order_by(VatFiling.created_at)).all()
    if user.role not in ("Admin", "Manager"):
        rows = [f for f in rows if f.staff_id == user.id]
    return [serialize(db, f) for f in rows]


@router.post("/filings/open")
def open_filing(body: OpenIn, user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Open (or return the existing open) filing for the duty's current period."""
    duty = get_scoped_or_404(db, Duty, body.duty_id, user)
    if duty.kind != "vat":
        raise conflict("The VAT Filing Engine only runs on VAT duties")
    if duty.closed:
        raise conflict("This duty is closed")
    if user.id != duty.staff_id and user.role not in ("Admin", "Manager"):
        raise conflict("Only the responsible staff member (or management) can open the filing")
    existing = db.scalar(select(VatFiling).where(VatFiling.duty_id == duty.id,
                                                 VatFiling.status != "complete"))
    if existing:
        return serialize(db, existing, detail=True)
    prof = _get_profile(db, user.tenant_id, duty.client_id)
    ps, pe, pps = derive_period(duty, prof)
    f = VatFiling(tenant_id=duty.tenant_id, duty_id=duty.id, client_id=duty.client_id,
                  staff_id=duty.staff_id, period_start=ps, period_end=pe, prev_period_start=pps,
                  status="ledgers_pending")
    db.add(f)
    db.flush()
    stagger_txt = (f", stagger {STAGGERS[prof.tax_period_stagger]} from profile v{prof.version}"
                   if prof and prof.tax_period_stagger else "")
    _log(db, f, user.id, f"VAT filing period opened: {_period_label(f)} (derived from the duty schedule"
                         f"{stagger_txt}, due {fmt_d(duty.next_due)}). Stage 1 — collect the VAT ledger. "
                         f"Invoices dated before {pps:%d %b %Y} fall out of the filing window (VAT rule).")
    if prof:
        active = [k for k in FLAG_KEYS if prof.flags.get(k, {}).get("value") in ("yes", "not_sure")]
        _log(db, f, None, f"VAT client profile v{prof.version} auto-applied "
                          f"({prof.business_category}; flags: {', '.join(active) or 'none'}) — "
                          f"compliance checks will be pre-applied at the computation.")
    db.commit()
    return serialize(db, f, detail=True)


@router.get("/filings/{fid}")
def get_filing(fid: uuid.UUID, user: User = Depends(current_user), db: Session = Depends(get_db)):
    return serialize(db, _get(db, fid, user), detail=True)


# ---------- stage 1 & 2: uploads ----------

def _replace_items(db: Session, f: VatFiling, source: str, rows: list[dict]):
    from sqlalchemy import delete
    q = delete(VatFilingItem).where(VatFilingItem.filing_id == f.id, VatFilingItem.source == source)
    if source == "invoice":  # a register re-upload never wipes confirmed AI-extracted rows
        q = q.where(VatFilingItem.origin == "register")
    db.execute(q)
    for r in rows:
        db.add(VatFilingItem(tenant_id=f.tenant_id, filing_id=f.id, source=source,
                             row_no=r["row_no"], invoice_no=r["invoice_no"],
                             invoice_no_norm=r["invoice_no_norm"], invoice_date=r["invoice_date"],
                             party=r["party"], trn=r["trn"], emirate=r["emirate"],
                             net=r["net"], vat=r["vat"], type_=r["type"], category=r["category"],
                             notes=r["notes"]))
    db.flush()  # the session doesn't autoflush — reconciliation must see these rows


@router.post("/filings/{fid}/ledger")
def upload_ledger(fid: uuid.UUID, file: UploadFile, user: User = Depends(current_user),
                  db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "ledgers_pending", "invoices_pending")
    rows = _parse_upload(file.file.read(), LEDGER_COLUMNS, is_ledger=True)
    file.file.seek(0)
    stored = _store_upload(db, user, "vat_filing", f.id, file)
    _replace_items(db, f, "ledger", rows)
    f.ledger_file = {"file_id": str(stored.id), "name": stored.name, "rows": len(rows), "at": iso(now())}
    f.status = "invoices_pending"
    out_n = sum(1 for r in rows if r["type"] == "Output")
    _log(db, f, user.id, f"VAT ledger uploaded: {stored.name} — {len(rows)} row(s) parsed "
                         f"({out_n} Output / {len(rows) - out_n} Input). Stage 2 — collect the invoice register.")
    db.commit()
    return serialize(db, f, detail=True)


@router.post("/filings/{fid}/invoices")
def upload_invoices(fid: uuid.UUID, file: UploadFile,
                    evidence: list[UploadFile] = FileParam(default=[]),
                    user: User = Depends(current_user), db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "invoices_pending", "reconciled")
    rows = _parse_upload(file.file.read(), REGISTER_COLUMNS, is_ledger=False)
    file.file.seek(0)
    stored = _store_upload(db, user, "vat_filing", f.id, file)
    pdfs = [_store_upload(db, user, "vat_filing", f.id, e) for e in evidence]
    _replace_items(db, f, "invoice", rows)
    f.invoice_file = {"file_id": str(stored.id), "name": stored.name, "rows": len(rows), "at": iso(now())}
    if pdfs:
        f.invoice_evidence = [*f.invoice_evidence,
                              *[{"file_id": str(p.id), "name": p.name, "size": p.size} for p in pdfs]]
    pdf_txt = f" · {len(pdfs)} invoice PDF(s) stored as evidence" if pdfs else ""
    _log(db, f, user.id, f"Invoice register uploaded: {stored.name} — {len(rows)} row(s) parsed{pdf_txt}.")
    _reconcile(db, f, user)
    db.commit()
    return serialize(db, f, detail=True)


# ---------- stage 2 alternative: AI invoice extraction (mandatory human review) ----------

VAT_EXTRACT_MODEL = "claude-sonnet-4-6"
EXTRACT_MEDIA_TYPES = {".pdf": "application/pdf", ".png": "image/png", ".jpg": "image/jpeg",
                       ".jpeg": "image/jpeg", ".webp": "image/webp", ".gif": "image/gif"}
EXTRACT_FIELDS = ("invoice_no", "invoice_date", "party", "emirate", "net_amount", "vat_amount", "currency")
EXTRACT_PROMPT = (
    "You are extracting fields from a SINGLE UAE tax invoice (the attached document). "
    "Return ONLY a JSON object — no prose, no markdown fences — with exactly these keys:\n"
    '{"invoice_no": {"value": string|null, "confidence": "low"|"high"},\n'
    ' "invoice_date": {"value": "YYYY-MM-DD"|null, "confidence": "low"|"high"},\n'
    ' "party": {"value": string|null, "confidence": "low"|"high"},\n'
    ' "emirate": {"value": string|null, "confidence": "low"|"high"},\n'
    ' "net_amount": {"value": number|null, "confidence": "low"|"high"},\n'
    ' "vat_amount": {"value": number|null, "confidence": "low"|"high"},\n'
    ' "currency": {"value": string|null, "confidence": "low"|"high"}}\n'
    "Rules: if a field is unreadable or absent, return null — NEVER guess. "
    "party = the counterparty the invoice is issued to (or by, if this is a purchase invoice). "
    f"emirate only if visible on the document, one of: {', '.join(EMIRATES)} — else null. "
    "net_amount = amount before VAT; vat_amount = the VAT charged. "
    "currency = the ISO code shown (AED, USD, …). Mark confidence low when the print is unclear, "
    "the value is inferred from layout, or multiple candidates exist."
)


def _extract_invoice(data: bytes, media_type: str) -> dict:
    """One document → extracted field dict. Raises on any failure (caller isolates per file)."""
    import base64
    import json as _json

    import anthropic

    key = get_settings().ANTHROPIC_API_KEY
    if not key:
        raise RuntimeError("ANTHROPIC_API_KEY is not configured")
    block_type = "document" if media_type == "application/pdf" else "image"
    client = anthropic.Anthropic(api_key=key)
    msg = client.messages.create(
        model=VAT_EXTRACT_MODEL, max_tokens=1024, temperature=0,
        messages=[{"role": "user", "content": [
            {"type": block_type, "source": {"type": "base64", "media_type": media_type,
                                            "data": base64.b64encode(data).decode()}},
            {"type": "text", "text": EXTRACT_PROMPT},
        ]}],
    )
    out = "".join(b.text for b in msg.content if b.type == "text").strip()
    if out.startswith("```"):
        out = out.strip("`")
        out = out[out.find("{"):out.rfind("}") + 1]
    parsed = _json.loads(out)
    return {k: (parsed.get(k) if isinstance(parsed.get(k), dict) else {"value": parsed.get(k), "confidence": "low"})
            for k in EXTRACT_FIELDS}


def _serialize_draft(d: VatExtractionDraft) -> dict:
    return {"id": d.id, "file_id": d.file_id, "file_name": d.file_name, "status": d.status,
            "fields": d.fields, "error": d.error, "created_at": d.created_at,
            "reviewed_by": d.reviewed_by, "reviewed_at": d.reviewed_at}


@router.post("/filings/{fid}/invoices/extract")
def extract_invoices(fid: uuid.UUID, files: list[UploadFile] = FileParam(...),
                     user: User = Depends(current_user), db: Session = Depends(get_db)):
    """AI extraction — an ALTERNATIVE input path at the invoice stage; mixes freely with
    the register upload. Multi-page PDFs are assumed to be ONE invoice per file. Every
    result is a DRAFT until a human reviews and confirms it."""
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "invoices_pending", "reconciled")
    max_files = int(os.getenv("VAT_EXTRACT_MAX_FILES", "25"))
    if len(files) > max_files:
        raise HTTPException(status_code=422, detail=f"{len(files)} files exceeds the per-batch extraction "
                            f"limit of {max_files} (VAT_EXTRACT_MAX_FILES) — split the batch to keep API "
                            f"costs predictable")
    results = []
    n_ok = 0
    for up in files:
        ext = "." + (up.filename or "").rsplit(".", 1)[-1].lower()
        data = up.file.read()
        up.file.seek(0)
        stored = _store_upload(db, user, "vat_filing", f.id, up)  # source kept as evidence
        draft = VatExtractionDraft(tenant_id=f.tenant_id, filing_id=f.id, file_id=stored.id,
                                   file_name=stored.name, created_by=user.id)
        media_type = EXTRACT_MEDIA_TYPES.get(ext)
        try:
            if media_type is None:
                raise RuntimeError(f"unsupported file type {ext!r} — PDF or image required")
            draft.fields = _extract_invoice(data, media_type)
            draft.status = "extracted"
            n_ok += 1
        except Exception as exc:
            draft.status = "failed"
            draft.error = f"extraction failed — enter manually ({type(exc).__name__})"
        db.add(draft)
        db.flush()
        results.append({"file_name": draft.file_name, "status": draft.status,
                        "draft_id": str(draft.id), "error": draft.error})
    _log(db, f, user.id, f"AI invoice extraction batch: {len(files)} file(s) → {n_ok} extracted, "
                         f"{len(files) - n_ok} failed. Drafts await human review — nothing enters the "
                         f"reconciliation until reviewed and confirmed.")
    db.commit()
    return {"results": results, "filing": serialize(db, f, detail=True)}


@router.post("/filings/{fid}/invoices/confirm-extracted")
def confirm_extractions(fid: uuid.UUID, body: ConfirmExtractionsIn,
                        user: User = Depends(current_user), db: Session = Depends(get_db)):
    """The mandatory review gate: staff-corrected rows become register items with
    origin=ai_extracted; reconciliation re-runs. Unconfirmed drafts never reconcile."""
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "invoices_pending", "reconciled")
    corrected = 0
    for row in body.rows:
        draft = db.scalar(select(VatExtractionDraft).where(VatExtractionDraft.filing_id == f.id,
                                                           VatExtractionDraft.id == row.draft_id))
        if draft is None:
            raise HTTPException(status_code=404, detail=f"extraction draft {row.draft_id} not found")
        if draft.status == "confirmed":
            raise conflict(f'"{draft.file_name}" is already confirmed')
        emirate = next((e for e in EMIRATES if e.lower() == row.emirate.strip().lower()), None)
        if emirate is None:
            raise HTTPException(status_code=422, detail=f"{draft.file_name}: emirate must be one of {', '.join(EMIRATES)}")
        try:
            inv_date = date.fromisoformat(row.invoice_date)
        except ValueError:
            raise HTTPException(status_code=422, detail=f"{draft.file_name}: invoice_date must be YYYY-MM-DD")
        currency = row.currency.strip().upper() or "AED"
        if currency != "AED" and not row.conversion_note.strip():
            raise conflict(f'"{draft.file_name}" is in {currency} — a manual-conversion note is mandatory '
                           f"(state the rate/source used to reach the AED amounts)")
        # count reviewer corrections vs the raw extraction
        raw = {k: (draft.fields.get(k) or {}).get("value") for k in EXTRACT_FIELDS}
        submitted = {"invoice_no": row.invoice_no.strip(), "invoice_date": row.invoice_date,
                     "party": row.party.strip(), "emirate": emirate,
                     "net_amount": row.net, "vat_amount": row.vat, "currency": currency}
        for k, v in submitted.items():
            old = raw.get(k)
            if k in ("net_amount", "vat_amount"):
                same = old is not None and abs(float(old) - float(v)) < 0.005
            else:
                same = old is not None and str(old).strip().lower() == str(v).strip().lower()
            if not same:
                corrected += 1
        note = f"AI-extracted from {draft.file_name}"
        if currency != "AED":
            note += f" · {currency} converted manually — {row.conversion_note.strip()}"
        db.add(VatFilingItem(tenant_id=f.tenant_id, filing_id=f.id, source="invoice",
                             origin="ai_extracted", row_no=0,
                             invoice_no=row.invoice_no.strip(),
                             invoice_no_norm=_norm_invoice_no(row.invoice_no),
                             invoice_date=inv_date, party=row.party.strip(), trn=None,
                             emirate=emirate, net=round(row.net, 2), vat=round(row.vat, 2),
                             type_=None, category="standard", notes=note))
        draft.status = "confirmed"
        draft.reviewed_by = user.id
        draft.reviewed_at = now()
    db.flush()
    _log(db, f, user.id, f"{len(body.rows)} invoice(s) extracted by AI, reviewed and confirmed by "
                         f"{user.name} — {corrected} field(s) corrected. Confirmed rows join the "
                         f"reconciliation as register items (origin: AI-extracted).")
    _reconcile(db, f, user)
    db.commit()
    return serialize(db, f, detail=True)


# ---------- stage 3: auto-reconciliation ----------

def _reconcile(db: Session, f: VatFiling, user: User):
    """Match key = normalized invoice_no + VAT amount (±0.01). Dates are NEVER used for
    matching — but rows dated before the PREVIOUS period's start are out of window."""
    items = db.scalars(select(VatFilingItem).where(VatFilingItem.filing_id == f.id)).all()
    for i in items:  # recon always restarts clean
        i.bucket = None
        i.resolution = None
        i.included = True

    in_window, out_of_window = [], []
    for i in items:
        if i.invoice_date < f.prev_period_start:
            i.bucket = "out_of_window"
            i.included = False
            out_of_window.append(i)
        else:
            in_window.append(i)

    # only Output (sales) ledger rows match against the client's issued-invoice register;
    # Input (purchase) rows have no register counterpart — they stay included, unbucketed
    ledger = [i for i in in_window if i.source == "ledger" and i.type_ == "Output"]
    invoices = [i for i in in_window if i.source == "invoice"]
    unused = list(invoices)
    matched = []
    for L in ledger:
        hit = next((I for I in unused
                    if I.invoice_no_norm == L.invoice_no_norm
                    and round(abs(float(I.vat) - float(L.vat)), 2) <= VAT_TOLERANCE), None)
        if hit is not None:
            L.bucket = hit.bucket = "matched"
            unused.remove(hit)
            matched.append((L, hit))
        else:
            L.bucket = "ledger_only"
    for I in unused:
        I.bucket = "invoice_only"

    ledger_only = [i for i in ledger if i.bucket == "ledger_only"]
    invoice_only = unused
    excel = _recon_workbook(f, items)
    fname = f"VAT Reconciliation {f.period_start:%b} - {f.period_end:%b %Y}.xlsx"
    stored = _store_bytes(db, user, "client", f.client_id or f.id, fname, excel)
    f.recon = {
        "at": iso(now()), "excel_file_id": str(stored.id), "excel_name": fname,
        "matched": len(matched), "ledger_only": len(ledger_only),
        "invoice_only": len(invoice_only), "out_of_window": len(out_of_window),
    }
    f.status = "reconciled"
    diffs = len(ledger_only) + len(invoice_only)
    _log(db, f, user.id,
         f"Auto-reconciliation run (key: invoice no + VAT amount ±{VAT_TOLERANCE:.2f}; dates never matched): "
         f"{len(matched)} matched · {len(ledger_only)} in ledger not in invoices · "
         f"{len(invoice_only)} in invoices not in ledger · {len(out_of_window)} out of window "
         f"(dated before {f.prev_period_start:%d %b %Y} — VAT rule). "
         f"Reconciliation workbook stored on the client registry: {fname}. "
         + ("Resolve every difference to unlock the computation." if diffs
            else "No differences — the computation is unlocked."))


ORIGIN_LABELS = {"register": "Register", "ai_extracted": "AI-extracted",
                 "added_at_recon": "Added at recon", "ledger_correction": "Ledger correction"}


def _recon_workbook(f: VatFiling, items) -> bytes:
    """Built from the CURRENT items — regenerates on download, reflecting every resolution
    (excluded / resolved / added rows) with its origin, plus the Ledger corrections sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    AMBER = PatternFill("solid", fgColor="FDE9C8")
    HEAD = Font(bold=True)
    COLS = ["Source", "Origin", "Row", "Invoice No", "Invoice Date", "Party", "Emirate", "Net", "VAT",
            "Type", "Bucket", "Resolution"]

    def origin_label(i):
        if i.source == "ledger" and i.origin == "register":
            return "Ledger"
        return ORIGIN_LABELS.get(i.origin, i.origin)

    def res_label(i):
        r = i.resolution or {}
        if r.get("action") == "excluded":
            return f"excluded — {r.get('reason', '')}"
        if r.get("action"):
            return r["action"]
        return ""

    def put_rows(ws, rows, fill=None):
        ws.append(COLS)
        for c in ws[1]:
            c.font = HEAD
        for i in rows:
            ws.append([i.source, origin_label(i), i.row_no, i.invoice_no,
                       i.invoice_date.strftime("%d/%m/%Y"), i.party,
                       i.emirate, float(i.net), float(i.vat), i.type_ or "", i.bucket or "",
                       res_label(i)])
            if fill:
                for c in ws[ws.max_row]:
                    c.fill = fill

    matched = [i for i in items if i.bucket == "matched"]
    diffs = [i for i in items if i.bucket in ("ledger_only", "invoice_only")]
    oow = [i for i in items if i.bucket == "out_of_window"]
    corrections = [i for i in items if i.origin == "ledger_correction"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["VAT Reconciliation"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Period: {_period_label(f)}"])
    ws.append([f"Window rule: invoices dated before {f.prev_period_start:%d %b %Y} are excluded (VAT rule)"])
    ws.append([])
    ws.append(["Matched", len(matched) // 2])
    ws.append(["In ledger, not in invoices", sum(1 for i in diffs if i.bucket == "ledger_only")])
    ws.append(["In invoices, not in ledger", sum(1 for i in diffs if i.bucket == "invoice_only")])
    ws.append(["Out of window (VAT rule)", len(oow)])
    ws.append(["Ledger corrections (to be booked by the client)", len(corrections)])
    put_rows(wb.create_sheet("Matched"), matched)
    put_rows(wb.create_sheet("Differences"), diffs, fill=AMBER)
    put_rows(wb.create_sheet("Excluded"), oow)
    cw = wb.create_sheet("Ledger corrections")
    cw.append(["Invoices added to the VAT workings that must also be booked in the client's ledger:"])
    cw["A1"].font = Font(italic=True, size=9)
    put_rows_start = cw.max_row
    cw.append(COLS + ["Correction note"])
    for c in cw[put_rows_start + 1]:
        c.font = HEAD
    for i in corrections:
        cw.append([i.source, origin_label(i), i.row_no, i.invoice_no,
                   i.invoice_date.strftime("%d/%m/%Y"), i.party, i.emirate, float(i.net),
                   float(i.vat), i.type_ or "", i.bucket or "", res_label(i),
                   (i.resolution or {}).get("correction_note") or i.notes or ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _refresh_recon_counts(db: Session, f: VatFiling):
    items = db.scalars(select(VatFilingItem).where(VatFilingItem.filing_id == f.id)).all()
    f.recon = {**(f.recon or {}),
               "matched": sum(1 for i in items if i.source == "ledger" and i.bucket == "matched"),
               "ledger_only": sum(1 for i in items if i.bucket == "ledger_only"),
               "invoice_only": sum(1 for i in items if i.bucket == "invoice_only"),
               "out_of_window": sum(1 for i in items if i.bucket == "out_of_window")}


def _item(db: Session, f: VatFiling, item_id: uuid.UUID) -> VatFilingItem:
    it = db.scalar(select(VatFilingItem).where(VatFilingItem.filing_id == f.id, VatFilingItem.id == item_id))
    if it is None:
        raise HTTPException(status_code=404, detail="filing item not found")
    return it


@router.post("/filings/{fid}/items/{item_id}/request-invoice")
def request_missing_invoice(fid: uuid.UUID, item_id: uuid.UUID, body: MailIn,
                            user: User = Depends(current_user), db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "reconciled")
    it = _item(db, f, item_id)
    if it.bucket not in ("ledger_only", "invoice_only"):
        raise conflict("Only reconciliation differences can be chased with the client")
    emails.send_client(str(body.to), body.subject, body.body)
    db.add(VatClientRequest(tenant_id=f.tenant_id, filing_id=f.id, kind="missing_invoice",
                            item_id=it.id, to_email=str(body.to), subject=body.subject, by_user=user.id))
    it.resolution = {"action": "requested", "by": str(user.id), "at": iso(now()), "to": str(body.to)}
    _log(db, f, user.id, f'Missing-invoice request emailed to {body.to} for "{it.invoice_no}" '
                         f'({it.party}, VAT {float(it.vat):,.2f}) — subject: "{body.subject}"')
    db.commit()
    return serialize(db, f, detail=True)


@router.post("/filings/{fid}/items/{item_id}/resolve")
def resolve_item(fid: uuid.UUID, item_id: uuid.UUID, user: User = Depends(current_user),
                 db: Session = Depends(get_db)):
    """The requested invoice arrived (or the ledger was corrected) — the row counts as resolved."""
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "reconciled")
    it = _item(db, f, item_id)
    if it.bucket not in ("ledger_only", "invoice_only"):
        raise conflict("Only reconciliation differences can be resolved")
    it.resolution = {**(it.resolution or {}), "action": "resolved", "by": str(user.id), "at": iso(now())}
    _log(db, f, user.id, f'Difference resolved: "{it.invoice_no}" ({it.source}) — supporting record received.')
    db.commit()
    return serialize(db, f, detail=True)


@router.post("/filings/{fid}/items/{item_id}/exclude")
def exclude_item(fid: uuid.UUID, item_id: uuid.UUID, body: ReasonIn,
                 user: User = Depends(current_user), db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "reconciled")
    it = _item(db, f, item_id)
    if it.bucket not in ("ledger_only", "invoice_only"):
        raise conflict("Only reconciliation differences can be excluded")
    it.resolution = {"action": "excluded", "reason": body.reason.strip(), "by": str(user.id), "at": iso(now())}
    it.included = False
    _log(db, f, user.id, f'Difference EXCLUDED from this filing: "{it.invoice_no}" ({it.source}) — '
                         f'mandatory reason: "{body.reason.strip()}"')
    db.commit()
    return serialize(db, f, detail=True)


def _targeted_match(it: VatFilingItem, new_item: VatFilingItem) -> bool:
    """Re-run matching for this pair only (same key as the full recon)."""
    if (new_item.invoice_no_norm == it.invoice_no_norm
            and round(abs(float(new_item.vat) - float(it.vat)), 2) <= VAT_TOLERANCE):
        it.bucket = new_item.bucket = "matched"
        it.resolution = None
        return True
    new_item.bucket = "invoice_only" if new_item.source == "invoice" else "ledger_only"
    return False


@router.post("/filings/{fid}/items/{item_id}/add-to-register")
def add_to_register(
    fid: uuid.UUID, item_id: uuid.UUID,
    invoice_no: str = Form(...), invoice_date: str = Form(...), party: str = Form(...),
    emirate: str = Form(...), net: float = Form(...), vat: float = Form(...),
    note: str = Form(""),
    evidence: list[UploadFile] = FileParam(default=[]),
    user: User = Depends(current_user), db: Session = Depends(get_db),
):
    """'Invoice found — add to register': resolves an in-ledger-not-in-invoices difference
    by adding the obtained invoice as a register item (origin=added_at_recon). The invoice
    file is REQUIRED evidence."""
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "reconciled")
    it = _item(db, f, item_id)
    if it.bucket != "ledger_only":
        raise conflict("Only in-ledger-not-in-invoices differences can be resolved this way")
    if not evidence:
        raise HTTPException(status_code=422, detail="The obtained invoice file (PDF/image) is required as evidence")
    matched_emirate = next((e for e in EMIRATES if e.lower() == emirate.strip().lower()), None)
    if matched_emirate is None:
        raise HTTPException(status_code=422, detail=f"emirate must be one of {', '.join(EMIRATES)}")
    try:
        inv_date = date.fromisoformat(invoice_date)
    except ValueError:
        raise HTTPException(status_code=422, detail="invoice_date must be YYYY-MM-DD")
    stored = [_store_upload(db, user, "vat_filing", f.id, e) for e in evidence]
    names = ", ".join(s.name for s in stored)
    new_item = VatFilingItem(
        tenant_id=f.tenant_id, filing_id=f.id, source="invoice", origin="added_at_recon", row_no=0,
        invoice_no=invoice_no.strip(), invoice_no_norm=_norm_invoice_no(invoice_no),
        invoice_date=inv_date, party=party.strip(), trn=None, emirate=matched_emirate,
        net=round(net, 2), vat=round(vat, 2), type_=None, category=it.category,
        notes=f"Added at reconciliation — evidence: {names}" + (f" · {note.strip()}" if note.strip() else ""))
    db.add(new_item)
    db.flush()
    if _targeted_match(it, new_item):
        _log(db, f, user.id, f"Difference resolved — invoice {invoice_no.strip()} obtained and added to "
                             f"register by {user.name}, evidence attached ({names}).")
    else:
        _log(db, f, user.id, f"Invoice {invoice_no.strip()} added to register by {user.name} (evidence: "
                             f"{names}) — but it does NOT match the ledger row (invoice no / VAT differ): "
                             f"a new difference is open.")
    _refresh_recon_counts(db, f)
    db.commit()
    return serialize(db, f, detail=True)


@router.post("/filings/{fid}/items/{item_id}/add-to-ledger")
def add_to_ledger(fid: uuid.UUID, item_id: uuid.UUID, body: AddToLedgerIn,
                  user: User = Depends(current_user), db: Session = Depends(get_db)):
    """'Missing from client ledger — add ledger entry': resolves an in-invoices-not-in-ledger
    difference with a ledger item (origin=ledger_correction) + MANDATORY correction note.
    The correction lands on the workbook's 'Ledger corrections' sheet and in the computation
    email so the client's books get fixed too."""
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "reconciled")
    it = _item(db, f, item_id)
    if it.bucket != "invoice_only":
        raise conflict("Only in-invoices-not-in-ledger differences can be resolved this way")
    if body.category not in SUPPLY_CATEGORIES:
        raise HTTPException(status_code=422, detail=f"category must be one of {sorted(SUPPLY_CATEGORIES)}")
    matched_emirate = next((e for e in EMIRATES if e.lower() == body.emirate.strip().lower()), None)
    if matched_emirate is None:
        raise HTTPException(status_code=422, detail=f"emirate must be one of {', '.join(EMIRATES)}")
    try:
        inv_date = date.fromisoformat(body.invoice_date)
    except ValueError:
        raise HTTPException(status_code=422, detail="invoice_date must be YYYY-MM-DD")
    new_item = VatFilingItem(
        tenant_id=f.tenant_id, filing_id=f.id, source="ledger", origin="ledger_correction", row_no=0,
        invoice_no=body.invoice_no.strip(), invoice_no_norm=_norm_invoice_no(body.invoice_no),
        invoice_date=inv_date, party=body.party.strip(), trn=None, emirate=matched_emirate,
        net=round(body.net, 2), vat=round(body.vat, 2), type_=body.type, category=body.category,
        notes=body.note.strip(),
        resolution={"action": "resolved", "via": "ledger_correction",
                    "correction_note": body.note.strip(), "by": str(user.id), "at": iso(now())})
    db.add(new_item)
    db.flush()
    if _targeted_match(it, new_item):
        _log(db, f, user.id, f"Difference resolved — missing ledger entry added by {user.name} for invoice "
                             f'{body.invoice_no.strip()} ({body.type}, {SUPPLY_CATEGORIES[body.category]}) — '
                             f'correction note: "{body.note.strip()}". The client must book this in their '
                             f"ledger (listed on the Ledger corrections sheet and in the computation email).")
    else:
        _log(db, f, user.id, f"Ledger entry {body.invoice_no.strip()} added by {user.name} — but it does "
                             f"NOT match the register row (invoice no / VAT differ): a new difference is open.")
    _refresh_recon_counts(db, f)
    db.commit()
    return serialize(db, f, detail=True)


@router.get("/filings/{fid}/recon-workbook")
def recon_workbook(fid: uuid.UUID, user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Regenerated on every download — reflects the current resolutions and their origins."""
    f = _get(db, fid, user)
    if not f.recon:
        raise conflict("No reconciliation has run yet")
    items = db.scalars(select(VatFilingItem).where(VatFilingItem.filing_id == f.id)
                       .order_by(VatFilingItem.source, VatFilingItem.row_no)).all()
    return _xlsx_response(_recon_workbook(f, items),
                          f.recon.get("excel_name") or "VAT Reconciliation.xlsx")


# ---------- client requests (stage 1 & 2, template attached) ----------

@router.post("/filings/{fid}/request-from-client")
def request_from_client(fid: uuid.UUID, body: RequestFromClientIn,
                        user: User = Depends(current_user), db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    if body.kind not in ("ledger", "invoices"):
        raise HTTPException(status_code=422, detail="kind must be ledger or invoices")
    template = _ledger_template_bytes() if body.kind == "ledger" else _register_template_bytes()
    tname = "VAT Ledger Template.xlsx" if body.kind == "ledger" else "Invoice Register Template.xlsx"
    stored = _store_bytes(db, user, "vat_filing", f.id, tname, template)
    link = _file_link(stored)
    emails.send_client(str(body.to), body.subject,
                       body.body + f"\n\nTemplate ({tname}) — download link (valid {blobs.LINK_TTL_MIN} minutes):\n{link}")
    db.add(VatClientRequest(tenant_id=f.tenant_id, filing_id=f.id, kind=body.kind,
                            to_email=str(body.to), subject=body.subject, by_user=user.id))
    label = "VAT ledger" if body.kind == "ledger" else "invoice register"
    _log(db, f, user.id, f'{label.capitalize()} requested from client at {body.to} — '
                         f'subject: "{body.subject}" (template attached).')
    db.commit()
    return serialize(db, f, detail=True)


# ---------- stage 4: computation ----------

def _included_ledger_rows(db: Session, f: VatFiling) -> list[VatFilingItem]:
    """Ledger rows entering the computation: matched Output rows, resolved differences,
    and Input rows (never register-matched) — excluded and out-of-window rows never."""
    items = db.scalars(select(VatFilingItem).where(VatFilingItem.filing_id == f.id,
                                                   VatFilingItem.source == "ledger")).all()
    out = []
    for i in items:
        if not i.included or i.bucket == "out_of_window":
            continue
        if i.bucket == "ledger_only" and (i.resolution or {}).get("action") != "resolved":
            continue
        out.append(i)
    return out


def _flag_value(profile: VatClientProfile | None, key: str) -> str | None:
    if profile is None:
        return None
    return (profile.flags.get(key) or {}).get("value", "no")


def _compliance_checks(profile: VatClientProfile | None, present: dict) -> list[dict]:
    """Profile × data rules. 'not_sure' is treated as Yes for warning purposes.
    kind=warning → needs an explicit proceed-despite-warning note;
    kind=confirmation → mandatory tick before the computation can be confirmed."""
    checks = []
    yes = lambda k: _flag_value(profile, k) in ("yes", "not_sure")  # noqa: E731
    labels = {"zero_rated": "zero-rated", "exempt": "exempt", "margin": "margin-scheme",
              "rcm_import": "RCM-import", "out_of_scope": "out-of-scope (designated zone)"}
    if profile is not None:
        if yes("has_zero_rated") and present["zero_rated"] == 0:
            checks.append({"id": "zero_rated_expected_missing", "kind": "warning",
                           "text": "The client profile expects zero-rated supplies (Art. 45 — exports, "
                                   "international transport…) but this period's ledger has none — confirm "
                                   "with the client that nothing was missed before filing."})
        for cat, flag in (("zero_rated", "has_zero_rated"), ("exempt", "has_exempt"),
                          ("margin", "margin_scheme"), ("rcm_import", "rcm_imports"),
                          ("out_of_scope", "designated_zone")):
            if present[cat] > 0 and _flag_value(profile, flag) == "no":
                checks.append({"id": f"{cat}_unexpected", "kind": "warning",
                               "text": f"The ledger contains {present[cat]} {labels[cat]} row(s) but the "
                                       f"client profile says No — update the profile or correct the ledger."})
        if yes("margin_scheme") and present["margin"] > 0:
            checks.append({"id": "margin_confirmation", "kind": "confirmation",
                           "text": f"Margin-scheme eligibility confirmed for the {present['margin']} "
                                   f"margin row(s) (Art. 29, Exec. Regs): goods previously subject to VAT, "
                                   f"purchased from non-registrants or under the scheme — VAT computed on "
                                   f"the profit margin (sale − purchase), not the full sale value."})
        if yes("rcm_imports"):
            checks.append({"id": "rcm_confirmation", "kind": "confirmation",
                           "text": "Reverse-charge on imports (Art. 48) is self-assessed — output VAT in "
                                   "Box 3/6 mechanics and recoverable input VAT as applicable; import VAT "
                                   "flows via the customs-linked TRN."})
        if yes("blocked_input_risk"):
            checks.append({"id": "blocked_input_review", "kind": "confirmation",
                           "text": "Blocked input categories in the client's spend (entertainment, motor "
                                   "vehicles with personal use) have been excluded from input VAT recovery."})
    if present["zero_rated"] > 0:
        checks.append({"id": "zero_rated_evidence", "kind": "confirmation",
                       "text": f"Export/zero-rating evidence is retained for the {present['zero_rated']} "
                               f"zero-rated row(s) — official and commercial evidence (customs exit + "
                               f"transport documents) within the 90-day rule (Art. 45). Without evidence "
                               f"the supply is standard-rated at 5%."})
    if present["exempt"] > 0:
        checks.append({"id": "exempt_apportionment", "kind": "confirmation",
                       "text": "Input VAT apportionment has been considered for the exempt supplies in "
                               "this period (Art. 46 — exempt suppliers cannot recover related input VAT)."})
    return checks


@router.post("/filings/{fid}/draft-computation")
def draft_computation(fid: uuid.UUID, user: User = Depends(current_user), db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "reconciled")
    items = db.scalars(select(VatFilingItem).where(VatFilingItem.filing_id == f.id)).all()
    unresolved = _unresolved_count(items)
    if unresolved:
        raise conflict(f"{unresolved} reconciliation difference(s) unresolved — every difference must be "
                       f"matched, requested→resolved, or excluded with a reason before the computation")
    rows = _included_ledger_rows(db, f)
    if not rows:
        raise conflict("No includable ledger rows — nothing to compute")

    # VAT201-shaped split: RCM rows self-assess both sides; zero-rated/exempt add no output
    # VAT; out-of-scope (designated zone) rows sit outside the return boxes entirely
    rcm_rows = [r for r in rows if r.category == "rcm_import"]
    oos = [r for r in rows if r.category == "out_of_scope"]
    non_rcm = [r for r in rows if r.category not in ("rcm_import", "out_of_scope")]
    output = [r for r in non_rcm if r.type_ == "Output"]
    input_ = [r for r in non_rcm if r.type_ == "Input"]
    std = [r for r in output if r.category == "standard"]
    zero = [r for r in output if r.category == "zero_rated"]
    exempt = [r for r in output if r.category == "exempt"]
    margin = [r for r in output if r.category == "margin"]

    ssum = lambda rs, attr: round(sum(float(getattr(r, attr)) for r in rs), 2)  # noqa: E731
    std_vat, margin_vat = ssum(std, "vat"), ssum(margin, "vat")
    rcm_vat = ssum(rcm_rows, "vat")
    output_vat = round(std_vat + margin_vat + rcm_vat, 2)
    input_vat = round(ssum(input_, "vat") + rcm_vat, 2)
    net = round(output_vat - input_vat, 2)
    per_emirate = {}
    for r in std:
        pe = per_emirate.setdefault(r.emirate, {"taxable_sales": 0.0, "output_vat": 0.0, "rows": 0})
        pe["taxable_sales"] = round(pe["taxable_sales"] + float(r.net), 2)
        pe["output_vat"] = round(pe["output_vat"] + float(r.vat), 2)
        pe["rows"] += 1

    present = {"zero_rated": len(zero), "exempt": len(exempt), "margin": len(margin),
               "rcm_import": len(rcm_rows), "out_of_scope": len(oos)}
    profile = _get_profile(db, f.tenant_id, f.client_id)
    checks = _compliance_checks(profile, present)
    # refresh the registry copy of the workbook — the stored version reflects final resolutions
    if f.recon:
        stored_x = _store_bytes(db, user, "client", f.client_id or f.id,
                                f.recon.get("excel_name") or "VAT Reconciliation.xlsx",
                                _recon_workbook(f, items))
        f.recon = {**f.recon, "excel_file_id": str(stored_x.id)}
    excluded = sum(1 for i in items if (i.resolution or {}).get("action") == "excluded")
    f.computation = {
        "period": _period_label(f), "at": iso(now()), "by": str(user.id),
        "profile_version": profile.version if profile else None,
        "output_vat": output_vat, "input_vat": input_vat, "net": abs(net),
        "position": "payable" if net >= 0 else "refundable",
        "taxable_sales": ssum(std, "net"),  # standard-rated sales
        "per_emirate": per_emirate,
        "zero_rated": {"sales": ssum(zero, "net"), "rows": len(zero)},
        "exempt": {"sales": ssum(exempt, "net"), "rows": len(exempt)},
        "margin": {"sales": ssum(margin, "net"), "output_vat": margin_vat, "rows": len(margin)},
        "rcm": {"output_vat": rcm_vat, "input_vat": rcm_vat, "rows": len(rcm_rows)},
        "out_of_scope": {"sales": ssum(oos, "net"), "rows": len(oos)},  # outside the return boxes
        "checks": checks,
        "counts": {"included": len(rows), "output_rows": len(output), "input_rows": len(input_),
                   "matched": sum(1 for i in items if i.bucket == "matched") // 2,
                   "excluded": excluded,
                   "out_of_window": sum(1 for i in items if i.bucket == "out_of_window")},
        "confirmed": False,
    }
    f.status = "computation_draft"
    n_warn = sum(1 for c in checks if c["kind"] == "warning")
    n_conf = sum(1 for c in checks if c["kind"] == "confirmation")
    _log(db, f, user.id, f"Computation auto-drafted from {len(rows)} included ledger row(s) "
                         f"(profile v{profile.version if profile else '—'} applied): "
                         f"standard {ssum(std, 'net'):,.2f} / zero-rated {ssum(zero, 'net'):,.2f} / "
                         f"exempt {ssum(exempt, 'net'):,.2f} · output VAT {output_vat:,.2f} · "
                         f"input VAT {input_vat:,.2f} · net {abs(net):,.2f} "
                         f"{f.computation['position'].upper()}. Compliance checks: {n_warn} warning(s), "
                         f"{n_conf} mandatory confirmation(s). Review and confirm.")
    db.commit()
    return serialize(db, f, detail=True)


def _computation_pdf(f: VatFiling, client_name: str) -> bytes:
    """Minimal clean single-page PDF (no external deps) with the computation."""
    c = f.computation
    zr, ex, mg, rcm = (c.get("zero_rated") or {}), (c.get("exempt") or {}), (c.get("margin") or {}), (c.get("rcm") or {})
    oos = c.get("out_of_scope") or {}
    lines = [
        "VAT Return Computation",
        f"Client: {client_name}",
        f"Period: {c['period']}",
        "",
        f"Standard-rated sales (net): AED {c['taxable_sales']:,.2f}",
        f"Zero-rated sales:           AED {zr.get('sales', 0):,.2f}",
        f"Exempt supplies:            AED {ex.get('sales', 0):,.2f}",
        f"Margin-scheme sales:        AED {mg.get('sales', 0):,.2f}  (VAT on margin: {mg.get('output_vat', 0):,.2f})",
        f"RCM self-assessed:          output {rcm.get('output_vat', 0):,.2f} / input {rcm.get('input_vat', 0):,.2f}",
        f"Out of scope (desig. zone): AED {oos.get('sales', 0):,.2f}  (outside the return boxes)",
        f"Output VAT (total):         AED {c['output_vat']:,.2f}",
        f"Input VAT (recoverable):    AED {c['input_vat']:,.2f}",
        f"NET VAT {c['position'].upper():<12}        AED {c['net']:,.2f}",
        "",
        "Standard-rated sales per emirate:",
    ]
    for em, v in c["per_emirate"].items():
        lines.append(f"  {em:<18} AED {v['taxable_sales']:,.2f}  (output VAT {v['output_vat']:,.2f}, "
                     f"{v['rows']} invoice(s))")
    k = c["counts"]
    lines += ["", f"Basis: {k['included']} ledger rows included ({k['output_rows']} output / "
                  f"{k['input_rows']} input) · {k['matched']} reconciled matches · "
                  f"{k['excluded']} excluded · {k['out_of_window']} out of window.",
              "", "Please review and confirm your approval so we can file at the FTA."]

    def esc(s):
        return s.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")

    content = "BT /F1 11 Tf 14 TL 56 790 Td\n"
    for i, ln in enumerate(lines):
        content += f"({esc(ln)}) Tj T*\n"
    content += "ET"
    content_b = content.encode("latin-1", "replace")
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>",
        b"<< /Length " + str(len(content_b)).encode() + b" >>\nstream\n" + content_b + b"\nendstream",
    ]
    out = io.BytesIO()
    out.write(b"%PDF-1.4\n")
    offsets = []
    for n, body in enumerate(objs, start=1):
        offsets.append(out.tell())
        out.write(f"{n} 0 obj\n".encode() + body + b"\nendobj\n")
    xref_at = out.tell()
    out.write(f"xref\n0 {len(objs) + 1}\n0000000000 65535 f \n".encode())
    for off in offsets:
        out.write(f"{off:010d} 00000 n \n".encode())
    out.write(f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\nstartxref\n{xref_at}\n%%EOF".encode())
    return out.getvalue()


@router.post("/filings/{fid}/confirm-computation")
def confirm_computation(fid: uuid.UUID, body: ConfirmComputationIn | None = None,
                        user: User = Depends(current_user), db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "computation_draft")
    body = body or ConfirmComputationIn()

    checks = f.computation.get("checks") or []
    required = [c for c in checks if c["kind"] == "confirmation"]
    missing = [c for c in required if c["id"] not in body.confirmations]
    if missing:
        raise conflict(f"{len(missing)} mandatory confirmation(s) unticked: "
                       + " · ".join(c["id"] for c in missing))
    warnings = [c for c in checks if c["kind"] == "warning"]
    if warnings and not body.warning_note.strip():
        raise conflict("Compliance warning(s) present — an explicit 'proceed despite warning' reason "
                       "note is mandatory")
    stamped = []
    for c in checks:
        if c["kind"] == "confirmation":
            stamped.append({**c, "ticked_by": str(user.id), "ticked_by_name": user.name,
                            "ticked_at": iso(now())})
        else:
            stamped.append({**c, "acknowledged_by": str(user.id), "acknowledged_by_name": user.name,
                            "acknowledged_at": iso(now())})

    client = db.get(Client, f.client_id) if f.client_id else None
    client_name = client.name if client else "Client"
    f.computation = {**f.computation, "checks": stamped,
                     "warning_note": body.warning_note.strip() or None}
    pdf = _computation_pdf(f, client_name)
    pname = f"VAT Computation {f.period_start:%b} - {f.period_end:%b %Y}.pdf"
    stored = _store_bytes(db, user, "client", f.client_id or f.id, pname, pdf)
    f.computation = {**f.computation, "confirmed": True, "confirmed_at": iso(now()),
                     "pdf_file_id": str(stored.id), "pdf_name": pname}
    f.status = "awaiting_client_approval"
    if required:
        _log(db, f, user.id, f"Compliance confirmations ticked by {user.name}: "
                             + "; ".join(c["id"] for c in required) + ".")
    if warnings:
        _log(db, f, user.id, f'{len(warnings)} compliance warning(s) acknowledged by {user.name} — '
                             f'proceed-despite-warning reason: "{body.warning_note.strip()}"')
    _log(db, f, user.id, f"Computation reviewed and CONFIRMED by staff. Rendered as {pname} "
                         f"(stored on the client registry). Stage 5 — send to the client for approval.")
    db.commit()
    return serialize(db, f, detail=True)


@router.post("/filings/{fid}/send-computation")
def send_computation(fid: uuid.UUID, body: MailIn, user: User = Depends(current_user),
                     db: Session = Depends(get_db)):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "awaiting_client_approval")
    pdf_id = (f.computation or {}).get("pdf_file_id")
    pdf_row = db.get(FileModel, uuid.UUID(pdf_id)) if pdf_id else None
    link = f"\n\nComputation ({f.computation['pdf_name']}) — download link (valid {blobs.LINK_TTL_MIN} minutes):\n" \
           f"{_file_link(pdf_row)}" if pdf_row else ""
    corrections = db.scalars(select(VatFilingItem).where(
        VatFilingItem.filing_id == f.id, VatFilingItem.origin == "ledger_correction")).all()
    corr_txt = ""
    if corrections:
        corr_txt = ("\n\nThe following invoices were added to the VAT workings and should be booked "
                    "in your ledger:\n" + "\n".join(
                        f"- {i.invoice_no} — {i.party}, net AED {float(i.net):,.2f}, "
                        f"VAT AED {float(i.vat):,.2f} ({(i.resolution or {}).get('correction_note') or i.notes})"
                        for i in corrections))
    emails.send_client(str(body.to), body.subject, body.body + corr_txt + link)
    db.add(VatClientRequest(tenant_id=f.tenant_id, filing_id=f.id, kind="computation",
                            to_email=str(body.to), subject=body.subject, by_user=user.id))
    _log(db, f, user.id, f'Computation emailed to the client at {body.to} — subject: "{body.subject}" '
                         f"(PDF attached). Awaiting client approval.")
    db.commit()
    return serialize(db, f, detail=True)


@router.post("/filings/{fid}/client-approval")
def record_client_approval(
    fid: uuid.UUID,
    basis: str = Form("evidence_upload"),
    note: str = Form(""),
    evidence: list[UploadFile] = FileParam(default=[]),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Evidence upload, or declared-with-basis + mandatory note (proposal-confirmation discipline)."""
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "awaiting_client_approval")
    if basis not in APPROVAL_BASES:
        raise HTTPException(status_code=422, detail=f"basis must be one of {sorted(APPROVAL_BASES)}")
    stored = [_store_upload(db, user, "client", f.client_id or f.id, e) for e in evidence]
    if basis == "evidence_upload" and not stored:
        raise HTTPException(status_code=422, detail="Upload the client's approval, or pick a declared basis")
    if basis != "evidence_upload" and not note.strip():
        raise conflict("A note describing exactly how the client approved is mandatory")
    f.client_approval = {"basis": basis, "label": APPROVAL_BASES[basis], "note": note.strip() or None,
                         "at": iso(now()), "by": str(user.id),
                         "evidence": [{"file_id": str(s.id), "name": s.name} for s in stored]}
    f.status = "ready_to_file"
    ev_txt = f' Evidence on file: {", ".join(s.name for s in stored)}.' if stored else ""
    _log(db, f, user.id, f"Client approval recorded — basis: {APPROVAL_BASES[basis]}"
                         f"{'; note: ' + repr(note.strip()) if note.strip() else ''}.{ev_txt} "
                         f"Stage 6 — file at the FTA.")
    db.commit()
    return serialize(db, f, detail=True)


# ---------- stage 6: file at FTA → complete the linked duty ----------

@router.post("/filings/{fid}/file-at-fta")
def file_at_fta(
    fid: uuid.UUID,
    note: str = Form(""),
    acknowledgement: list[UploadFile] = FileParam(default=[]),
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    f = _get(db, fid, user)
    _require_staff_open(f, user)
    _require_status(f, "ready_to_file")
    if not acknowledgement:
        raise HTTPException(status_code=422, detail="The FTA acknowledgement upload is required")
    duty = db.get(Duty, f.duty_id)
    if duty is None or duty.closed:
        raise conflict("The linked duty is missing or closed")
    if duty.staff_id != user.id:
        raise conflict("Only the responsible staff member can complete this duty")

    stored = [_store_upload(db, user, "client", f.client_id or f.id, a) for a in acknowledgement]
    c = f.computation
    record = {
        "period": c["period"],
        "position": c["position"],
        "net VAT (AED)": f"{c['net']:,.2f}",
        "output VAT (AED)": f"{c['output_vat']:,.2f}",
        "input VAT (AED)": f"{c['input_vat']:,.2f}",
        "zero-rated sales (AED)": f"{(c.get('zero_rated') or {}).get('sales', 0):,.2f}",
        "exempt sales (AED)": f"{(c.get('exempt') or {}).get('sales', 0):,.2f}",
        "taxable sales per emirate": "; ".join(
            f"{em} {v['taxable_sales']:,.2f}" for em, v in c["per_emirate"].items()),
    }
    # the module's ONE outward touch: complete the linked duty through the existing machinery
    apply_completion(db, duty, user, method="proof", stored=stored, record_obj=record,
                     note=note or f"Filed via VAT Filing Engine — {c['period']}")

    f.fta_ack = {"at": iso(now()), "by": str(user.id), "note": note or None,
                 "evidence": [{"file_id": str(s.id), "name": s.name} for s in stored]}
    f.status = "complete"
    f.completed_at = now()
    _log(db, f, user.id, f"FILED AT FTA — acknowledgement on file: {', '.join(s.name for s in stored)}. "
                         f"FILING COMPLETE — {c['period']}, net {c['net']:,.2f} {c['position'].upper()}. "
                         f"Linked duty completed with method=proof (record pre-filled from the computation); "
                         f"the schedule rolls forward automatically. Trail sealed.")
    db.commit()
    return {"filing": serialize(db, f, detail=True),
            "duty": {"id": duty.id, "next_due": duty.next_due, "closed": duty.closed}}
