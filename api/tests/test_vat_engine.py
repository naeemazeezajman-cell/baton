"""VAT Filing Engine — separate, removable module. Template parsing, reconciliation
buckets (tolerance + window rule), the computation gate, computation math, duty
completion pre-fill, the env flag, and tenancy."""

import io
import json
from datetime import date

from sqlalchemy import text as sql

from app.db import engine
from .conftest import bootstrap_tenant, login_after_reset
from .test_onboarding import setup_firm

H = lambda ctx, who: ctx[who]["headers"]  # noqa: E731

LEDGER_ROWS = [
    # Invoice No, Invoice Date, Party Name, TRN, Emirate, Net, VAT, Type
    ["INV-001", date(2026, 5, 10), "Acme LLC", "TRN-1", "Dubai", 1000, 50, "Output"],
    ["INV-002", date(2026, 6, 15), "Beta FZE", "TRN-2", "Sharjah", 2000, 100, "Output"],
    ["INV-003", date(2026, 7, 1), "Gamma DMCC", "TRN-3", "Dubai", 3000, 150, "Output"],
    ["INV-004", date(2026, 5, 20), "Delta Est", "TRN-4", "Ajman", 500, 25, "Output"],
    ["PUR-001", date(2026, 6, 1), "Supplier Co", "TRN-5", "Dubai", 800, 40, "Input"],
    ["OLD-001", date(2026, 1, 15), "Ancient LLC", "TRN-6", "Dubai", 999, 49.95, "Output"],
]
REGISTER_ROWS = [
    # Invoice No, Invoice Date, Party, Emirate, Net, VAT, Notes
    ["INV-001", date(2026, 5, 10), "Acme LLC", "Dubai", 1000, 50, ""],
    ["inv-002 ", date(2026, 6, 16), "Beta FZE", "Sharjah", 2000, 100.01, "case+space+tolerance; date differs — never matched on"],
    ["INV-003", date(2026, 7, 1), "Gamma DMCC", "Dubai", 3000, 151, "VAT differs beyond tolerance"],
    ["INV-900", date(2026, 6, 20), "Mystery Co", "Dubai", 700, 35, "not in ledger"],
    ["OLD-900", date(2026, 1, 2), "Ancient LLC", "Dubai", 400, 20, "out of window"],
]


def make_vat_duty(client, ctx, next_due="2026-08-28T00:00:00Z", cadence="quarterly", client_id=None):
    r = client.post("/duties", json={
        "staff_id": ctx["staff"]["id"], "client_name": "Gulf Horizon Trading LLC",
        "client_id": client_id,
        "service": "VAT Filing", "cadence": cadence, "next_due": next_due,
        "contact": {"name": "Mariam", "email": "accounts@gulfhorizon.ae"},
    }, headers=H(ctx, "manager"))
    assert r.status_code == 201, r.text
    return r.json()


def make_client(name="Gulf Horizon Trading LLC"):
    """Direct client row (skips the proposal flow — tests only)."""
    with engine.begin() as conn:
        tid = conn.execute(sql("SELECT tenant_id FROM users LIMIT 1")).scalar()
        cid = conn.execute(sql(
            "INSERT INTO clients (tenant_id, ref, name, contact) "
            "VALUES (:t, 'CL-001', :n, cast(:c AS jsonb)) RETURNING id"
        ), {"t": tid, "n": name,
            "c": json.dumps({"email": "accounts@gulfhorizon.ae", "contactPerson": "Mariam"})}).scalar()
    return str(cid)


def flags(**kw):
    return {k: {"value": v} for k, v in kw.items()}


def open_filing(client, ctx, duty_id):
    r = client.post("/vat-engine/filings/open", json={"duty_id": duty_id}, headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    return r.json()


def get_template(client, ctx, which):
    r = client.get(f"/vat-engine/templates/{which}", headers=H(ctx, "staff"))
    assert r.status_code == 200
    assert r.content[:2] == b"PK"  # xlsx = zip container
    return r.content


def fill_template(template_bytes, rows):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_ledger(client, ctx, fid, xlsx, expect=200):
    r = client.post(f"/vat-engine/filings/{fid}/ledger",
                    files={"file": ("Ledger.xlsx", xlsx,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    headers=H(ctx, "staff"))
    assert r.status_code == expect, r.text
    return r.json()


def upload_register(client, ctx, fid, xlsx, expect=200, pdfs=()):
    files = [("file", ("Register.xlsx", xlsx,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))]
    for name in pdfs:
        files.append(("evidence", (name, b"%PDF-1.4 evidence", "application/pdf")))
    r = client.post(f"/vat-engine/filings/{fid}/invoices", files=files, headers=H(ctx, "staff"))
    assert r.status_code == expect, r.text
    return r.json()


def drive_to_reconciled(client, ctx):
    duty = make_vat_duty(client, ctx)
    f = open_filing(client, ctx, duty["id"])
    ledger = fill_template(get_template(client, ctx, "ledger"), LEDGER_ROWS)
    upload_ledger(client, ctx, f["id"], ledger)
    register = fill_template(get_template(client, ctx, "invoice-register"), REGISTER_ROWS)
    f = upload_register(client, ctx, f["id"], register, pdfs=["inv-001.pdf"])
    return duty, f


def resolve_all_differences(client, ctx, f):
    items = {(i["source"], i["invoice_no"]): i for i in f["items"]}
    # INV-003 ledger side: request from client → mark resolved (include in filing)
    lid = items[("ledger", "INV-003")]["id"]
    r = client.post(f"/vat-engine/filings/{f['id']}/items/{lid}/request-invoice",
                    json={"to": "accounts@gulfhorizon.ae", "subject": "Missing invoice INV-003",
                          "body": "Please send INV-003."}, headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    r = client.post(f"/vat-engine/filings/{f['id']}/items/{lid}/resolve", headers=H(ctx, "staff"))
    assert r.status_code == 200
    # the register-side INV-003 row, INV-004 and INV-900 are excluded with reasons
    for key in (("invoice", "INV-003"), ("ledger", "INV-004"), ("invoice", "INV-900")):
        iid = items[key]["id"]
        r = client.post(f"/vat-engine/filings/{f['id']}/items/{iid}/exclude",
                        json={"reason": f"Not part of this filing — {key[1]}"}, headers=H(ctx, "staff"))
        assert r.status_code == 200, r.text
    return client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()


def test_period_derivation_staggered_quarter(client):
    ctx = setup_firm(client)
    duty = make_vat_duty(client, ctx, next_due="2026-08-28T00:00:00Z")  # staggered quarter
    f = open_filing(client, ctx, duty["id"])
    assert f["period_start"] == "2026-05-01" and f["period_end"] == "2026-07-31"
    assert f["prev_period_start"] == "2026-02-01"
    assert f["status"] == "ledgers_pending" and f["holder"] == ctx["staff"]["id"]
    # opening again returns the same filing, not a duplicate
    again = open_filing(client, ctx, duty["id"])
    assert again["id"] == f["id"]


def test_template_parse_happy_and_malformed(client):
    ctx = setup_firm(client)
    duty = make_vat_duty(client, ctx)
    f = open_filing(client, ctx, duty["id"])
    tpl = get_template(client, ctx, "ledger")

    # malformed: wrong columns entirely
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.append(["whatever"])
    wb.active.append(["Wrong", "Columns", "Here"])
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(f"/vat-engine/filings/{f['id']}/ledger",
                    files={"file": ("bad.xlsx", buf.getvalue(), "application/octet-stream")},
                    headers=H(ctx, "staff"))
    assert r.status_code == 422
    assert "Columns don't match the template" in r.json()["detail"]["reason"]

    # malformed rows: bad date, bad type, non-numeric amount → row-level errors, nothing stored
    bad = fill_template(tpl, [
        ["INV-A", "not-a-date", "Party", "T", "Dubai", 100, 5, "Output"],
        ["INV-B", date(2026, 6, 1), "Party", "T", "Dubai", 100, 5, "Sideways"],
        ["INV-C", date(2026, 6, 1), "Party", "T", "Atlantis", 100, 5, "Output"],
        ["INV-D", date(2026, 6, 1), "Party", "T", "Dubai", "lots", 5, "Output"],
    ])
    r = client.post(f"/vat-engine/filings/{f['id']}/ledger",
                    files={"file": ("bad2.xlsx", bad, "application/octet-stream")}, headers=H(ctx, "staff"))
    assert r.status_code == 422
    errors = r.json()["detail"]["errors"]
    assert len(errors) == 4
    assert any("Row 3" in e and "not a valid date" in e for e in errors)
    assert any("Row 4" in e and "Type must be Output or Input" in e for e in errors)
    assert any("Row 5" in e and "Emirate" in e for e in errors)
    assert any("Row 6" in e and "must be numbers" in e for e in errors)
    detail = client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()
    assert detail["status"] == "ledgers_pending" and detail["items"] == []

    # happy: parses, advances, counts Output/Input
    good = fill_template(tpl, LEDGER_ROWS)
    out = upload_ledger(client, ctx, f["id"], good)
    assert out["status"] == "invoices_pending"
    assert out["ledger_file"]["rows"] == len(LEDGER_ROWS)
    assert sum(1 for i in out["items"] if i["type"] == "Input") == 1


def test_recon_buckets_tolerance_and_window_rule(client):
    ctx = setup_firm(client)
    _, f = drive_to_reconciled(client, ctx)
    assert f["status"] == "reconciled"
    # matched: INV-001 exact; INV-002 via normalization + ±0.01 tolerance (dates differ — never matched on)
    assert f["recon"]["matched"] == 2
    # ledger-only: INV-003 (VAT differs beyond tolerance) + INV-004 (absent from register)
    assert f["recon"]["ledger_only"] == 2
    # invoice-only: register INV-003 (unmatched counterpart) + INV-900
    assert f["recon"]["invoice_only"] == 2
    # out of window: OLD-001 (ledger) + OLD-900 (register), dated before 2026-02-01
    assert f["recon"]["out_of_window"] == 2
    by = {(i["source"], i["invoice_no"]): i for i in f["items"]}
    assert by[("ledger", "INV-002")]["bucket"] == "matched"
    assert by[("ledger", "OLD-001")]["bucket"] == "out_of_window" and not by[("ledger", "OLD-001")]["included"]
    assert by[("invoice", "OLD-900")]["bucket"] == "out_of_window"
    # Input (purchase) rows never register-match — no bucket, still included
    assert by[("ledger", "PUR-001")]["bucket"] is None and by[("ledger", "PUR-001")]["included"]
    # the reconciliation workbook exists and is a valid xlsx with the four sheets
    import io as _io
    from openpyxl import load_workbook
    from app.db import engine
    from sqlalchemy import text as sql
    with engine.begin() as conn:
        blob_path = conn.execute(sql("SELECT blob_path FROM files WHERE id = :i"),
                                 {"i": f["recon"]["excel_file_id"]}).scalar()
    from app import blobs
    wb = load_workbook(_io.BytesIO(blobs.read_blob(blob_path)))
    assert wb.sheetnames == ["Summary", "Matched", "Differences", "Excluded", "Ledger corrections"]


def test_computation_gate_blocks_until_resolved(client):
    ctx = setup_firm(client)
    _, f = drive_to_reconciled(client, ctx)
    r = client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 409
    assert "unresolved" in r.json()["detail"]["reason"]
    f = resolve_all_differences(client, ctx, f)
    assert f["unresolved_differences"] == 0
    r = client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "computation_draft"


def test_computation_math_from_fixture(client):
    ctx = setup_firm(client)
    _, f = drive_to_reconciled(client, ctx)
    f = resolve_all_differences(client, ctx, f)
    c = client.post(f"/vat-engine/filings/{f['id']}/draft-computation",
                    headers=H(ctx, "staff")).json()["computation"]
    # included: INV-001 (1000/50 Dubai) + INV-002 (2000/100 Sharjah) + resolved INV-003
    # (3000/150 Dubai) + Input PUR-001 (800/40). Excluded INV-004; out-of-window OLD-001.
    assert c["output_vat"] == 300.0
    assert c["input_vat"] == 40.0
    assert c["net"] == 260.0 and c["position"] == "payable"
    assert c["taxable_sales"] == 6000.0
    assert c["per_emirate"]["Dubai"] == {"taxable_sales": 4000.0, "output_vat": 200.0, "rows": 2}
    assert c["per_emirate"]["Sharjah"] == {"taxable_sales": 2000.0, "output_vat": 100.0, "rows": 1}
    assert c["counts"]["included"] == 4 and c["counts"]["output_rows"] == 3 and c["counts"]["input_rows"] == 1
    assert c["counts"]["matched"] == 2 and c["counts"]["excluded"] == 3 and c["counts"]["out_of_window"] == 2
    assert c["period"] == "01 May 2026 – 31 Jul 2026"


def drive_to_complete(client, ctx):
    duty, f = drive_to_reconciled(client, ctx)
    f = resolve_all_differences(client, ctx, f)
    client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200 and r.json()["status"] == "awaiting_client_approval"
    r = client.post(f"/vat-engine/filings/{f['id']}/send-computation",
                    json={"to": "accounts@gulfhorizon.ae", "subject": "VAT computation for approval",
                          "body": "Please approve."}, headers=H(ctx, "staff"))
    assert r.status_code == 200
    r = client.post(f"/vat-engine/filings/{f['id']}/client-approval",
                    data={"basis": "email_approval", "note": "Approved by Mariam by email 11 Jul"},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200 and r.json()["status"] == "ready_to_file"
    r = client.post(f"/vat-engine/filings/{f['id']}/file-at-fta",
                    data={"note": "Filed on portal"},
                    files=[("acknowledgement", ("FTA-ack.pdf", b"%PDF ack", "application/pdf"))],
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    return duty, r.json()


def test_completion_prefills_duty_record_and_rolls_schedule(client):
    ctx = setup_firm(client)
    duty, out = drive_to_complete(client, ctx)
    assert out["filing"]["status"] == "complete"
    # duty completed through the existing machinery with method=proof + pre-filled record
    duties = client.get("/duties", headers=H(ctx, "staff")).json()
    d = next(x for x in duties if x["id"] == duty["id"])
    assert d["next_due"].startswith("2026-11-28")  # quarterly roll from 28 Aug
    comp = d["history"][-1]
    assert comp["method"] == "proof"
    rec = comp["record"]
    assert rec["period"] == "01 May 2026 – 31 Jul 2026"
    assert rec["position"] == "payable"
    assert rec["net VAT (AED)"] == "260.00"
    assert rec["output VAT (AED)"] == "300.00" and rec["input VAT (AED)"] == "40.00"
    assert rec["zero-rated sales (AED)"] == "0.00" and rec["exempt sales (AED)"] == "0.00"
    assert "Dubai 4,000.00" in rec["taxable sales per emirate"]
    assert "Sharjah 2,000.00" in rec["taxable sales per emirate"]
    assert comp["evidence"][0]["name"] == "FTA-ack.pdf"
    assert any("Trail sealed" in e["text"] for e in out["filing"]["events"])


def test_sealed_filing_blocks_mutations(client):
    ctx = setup_firm(client)
    _, out = drive_to_complete(client, ctx)
    fid = out["filing"]["id"]
    item_id = out["filing"]["items"][0]["id"]
    tpl = fill_template(get_template(client, ctx, "ledger"), LEDGER_ROWS)
    assert upload_ledger(client, ctx, fid, tpl, expect=409)
    r = client.post(f"/vat-engine/filings/{fid}/items/{item_id}/exclude", json={"reason": "x"},
                    headers=H(ctx, "staff"))
    assert r.status_code == 409
    r = client.post(f"/vat-engine/filings/{fid}/file-at-fta", data={"note": ""},
                    files=[("acknowledgement", ("a.pdf", b"%PDF", "application/pdf"))],
                    headers=H(ctx, "staff"))
    assert r.status_code == 409
    # read stays open
    assert client.get(f"/vat-engine/filings/{fid}", headers=H(ctx, "staff")).status_code == 200


def test_flag_off_404s_everything(client, monkeypatch):
    # enabled by default: unauthenticated hits the auth wall, not a 404
    assert client.get("/vat-engine/status").status_code == 401
    monkeypatch.setenv("VAT_ENGINE_ENABLED", "false")
    assert client.get("/vat-engine/status").status_code == 404
    assert client.get("/vat-engine/templates/ledger").status_code == 404
    assert client.get("/vat-engine/filings").status_code == 404
    assert client.post("/vat-engine/filings/open", json={}).status_code == 404
    assert client.post("/vat-engine/filings/00000000-0000-0000-0000-000000000000/invoices/extract",
                       files=[("files", ("x.pdf", b"%PDF", "application/pdf"))]).status_code == 404
    monkeypatch.delenv("VAT_ENGINE_ENABLED")
    assert client.get("/vat-engine/status").status_code == 401  # back on


# ---------- client VAT profile, wizard gate, compliance checks ----------

CAT_LEDGER = [
    ["S-1", date(2026, 5, 5), "Std Co", "T1", "Dubai", 1000, 50, "Output", "Standard (5%)"],
    ["S-2", date(2026, 5, 6), "Std Co", "T1", "Sharjah", 2000, 100, "Output", ""],  # blank → Standard
    ["Z-1", date(2026, 5, 7), "Exporter", "T2", "Dubai", 5000, 0, "Output", "Zero-rated (0%)"],
    ["E-1", date(2026, 5, 8), "Landlord", "T3", "Dubai", 3000, 0, "Output", "Exempt"],
    ["M-1", date(2026, 5, 9), "Car Dealer", "T4", "Dubai", 45000, 250, "Output", "Margin scheme"],
    ["R-1", date(2026, 5, 10), "US Vendor", "T5", "Dubai", 2000, 100, "Input", "RCM-Import"],
    ["P-1", date(2026, 5, 11), "Supplier", "T6", "Dubai", 800, 40, "Input", "Standard (5%)"],
]
CAT_REGISTER = [
    ["S-1", date(2026, 5, 5), "Std Co", "Dubai", 1000, 50, "", ""],
    ["S-2", date(2026, 5, 6), "Std Co", "Sharjah", 2000, 100, "", ""],
    ["Z-1", date(2026, 5, 7), "Exporter", "Dubai", 5000, 0, "", "Zero-rated (0%)"],
    ["E-1", date(2026, 5, 8), "Landlord", "Dubai", 3000, 0, "", "Exempt"],
    ["M-1", date(2026, 5, 9), "Car Dealer", "Dubai", 45000, 250, "", "Margin scheme"],
]

ALL_YES = flags(has_zero_rated="yes", has_exempt="yes", margin_scheme="yes", rcm_imports="yes")


def setup_client_filing(client, ctx, ledger=None, register=None, profile_flags=None):
    """Client row + linked VAT duty + open filing; optional profile + uploads through recon."""
    cid = make_client()
    duty = make_vat_duty(client, ctx, client_id=cid)
    if profile_flags is not None:
        r = client.post(f"/vat-engine/clients/{cid}/profile",
                        json={"nature_of_business": "General trading and used vehicles",
                              "business_category": "Trading", "flags": profile_flags},
                        headers=H(ctx, "staff"))
        assert r.status_code == 201, r.text
    f = open_filing(client, ctx, duty["id"])
    if ledger is not None:
        upload_ledger(client, ctx, f["id"], fill_template(get_template(client, ctx, "ledger"), ledger))
    if register is not None:
        f = upload_register(client, ctx, f["id"],
                            fill_template(get_template(client, ctx, "invoice-register"), register))
    return cid, duty, client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()


def test_wizard_gate_and_profile_change_events(client):
    ctx = setup_firm(client)
    cid = make_client()
    duty = make_vat_duty(client, ctx, client_id=cid)
    # no profile yet: 404 on GET, filing detail carries profile: null → frontend shows the wizard
    assert client.get(f"/vat-engine/clients/{cid}/profile", headers=H(ctx, "staff")).status_code == 404
    f = open_filing(client, ctx, duty["id"])
    assert f["profile"] is None

    # staff (assigned VAT duty) creates the profile; the open filing logs it as applied
    r = client.post(f"/vat-engine/clients/{cid}/profile",
                    json={"nature_of_business": "Trading in electronics",
                          "business_category": "Trading",
                          "flags": flags(has_zero_rated="not_sure")},
                    headers=H(ctx, "staff"))
    assert r.status_code == 201, r.text
    p = r.json()
    assert p["version"] == 1
    # "Not sure" is stored verbatim (renders the amber confirm-with-client chip)
    assert p["flags"]["has_zero_rated"]["value"] == "not_sure"
    # v1 history entry records every answer, structured
    v1 = p["updated"][0]
    assert v1["version"] == 1 and v1["by_name"] == "Priya Nair"
    by_field = {c["field"]: c for c in v1["changes"]}
    assert by_field["has_zero_rated"] == {"field": "has_zero_rated", "old": None,
                                          "new": "Not sure", "note": None}
    assert by_field["open_fta_matters"]["new"] == "No"
    f = client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()
    assert f["profile"]["version"] == 1  # returning visits skip straight to periods
    assert any("VAT client profile v1 recorded" in e["text"] and "Applied to this filing" in e["text"]
               for e in f["events"])
    # accountant may not edit; duplicate create refused
    assert client.post(f"/vat-engine/clients/{cid}/profile",
                       json={"business_category": "Trading", "flags": {}},
                       headers=H(ctx, "accountant")).status_code == 409
    assert client.post(f"/vat-engine/clients/{cid}/profile",
                       json={"business_category": "Trading", "flags": {}},
                       headers=H(ctx, "staff")).status_code == 409

    # edit: has_exempt No → Yes with a note → version bump, updated log, vat event
    r = client.patch(f"/vat-engine/clients/{cid}/profile",
                     json={"nature_of_business": "Trading in electronics",
                           "business_category": "Trading",
                           "flags": flags(has_zero_rated="not_sure",
                                          has_exempt="yes") | {"has_exempt": {"value": "yes", "note": "Now leasing residential units"}}},
                     headers=H(ctx, "manager"))
    assert r.status_code == 200, r.text
    p = r.json()
    assert p["version"] == 2
    # structured version row: {field, old → new, note}
    row = next(c for c in p["updated"][-1]["changes"] if c["field"] == "has_exempt")
    assert row == {"field": "has_exempt", "old": "No", "new": "Yes",
                   "note": "Now leasing residential units"}
    f = client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()
    assert any('Profile updated to v2' in e["text"] and 'has_exempt No → Yes' in e["text"]
               for e in f["events"])


def test_supply_category_parsing(client):
    ctx = setup_firm(client)
    _, _, f = setup_client_filing(client, ctx, ledger=CAT_LEDGER)
    cats = {i["invoice_no"]: i["category"] for i in f["items"] if i["source"] == "ledger"}
    assert cats == {"S-1": "standard", "S-2": "standard", "Z-1": "zero_rated", "E-1": "exempt",
                    "M-1": "margin", "R-1": "rcm_import", "P-1": "standard"}
    # invalid category value → row-level hard fail
    bad = fill_template(get_template(client, ctx, "ledger"),
                        [["X-1", date(2026, 5, 5), "P", "T", "Dubai", 100, 5, "Output", "Luxury rate"]])
    r = client.post(f"/vat-engine/filings/{f['id']}/ledger",
                    files={"file": ("bad.xlsx", bad, "application/octet-stream")}, headers=H(ctx, "staff"))
    assert r.status_code == 422
    assert any("Supply Category 'Luxury rate'" in e for e in r.json()["detail"]["errors"])


def test_compliance_rules_fire_both_directions(client):
    ctx = setup_firm(client)
    # direction 1: profile expects zero-rated, ledger has none → warning; note required to proceed
    _, _, f = setup_client_filing(client, ctx, ledger=LEDGER_ROWS[:2] + [LEDGER_ROWS[4]],
                                  register=REGISTER_ROWS[:2],
                                  profile_flags=flags(has_zero_rated="yes"))
    r = client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    checks = r.json()["computation"]["checks"]
    assert [c["id"] for c in checks] == ["zero_rated_expected_missing"]
    assert checks[0]["kind"] == "warning"
    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation", json={}, headers=H(ctx, "staff"))
    assert r.status_code == 409 and "warning" in r.json()["detail"]["reason"].lower()
    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation",
                    json={"warning_note": "Client confirmed no exports this quarter"}, headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    f2 = r.json()
    assert any('warning(s) acknowledged by Priya Nair' in e["text"]
               and "no exports this quarter" in e["text"] for e in f2["events"])
    assert f2["computation"]["checks"][0]["acknowledged_by_name"] == "Priya Nair"


def test_compliance_rules_unexpected_rows_and_mandatory_ticks(client):
    ctx = setup_firm(client)
    # direction 2: ledger contains exempt + margin rows but the profile says No to both
    _, _, f = setup_client_filing(client, ctx, ledger=CAT_LEDGER, register=CAT_REGISTER,
                                  profile_flags=flags(has_zero_rated="yes"))  # exempt/margin/rcm = No
    r = client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    checks = {c["id"]: c for c in r.json()["computation"]["checks"]}
    assert checks["exempt_unexpected"]["kind"] == "warning"
    assert checks["margin_unexpected"]["kind"] == "warning"
    assert checks["rcm_import_unexpected"]["kind"] == "warning"
    assert checks["exempt_apportionment"]["kind"] == "confirmation"
    assert checks["zero_rated_evidence"]["kind"] == "confirmation"  # export-evidence tick (Art. 45)
    assert "90-day rule" in checks["zero_rated_evidence"]["text"]
    assert "zero_rated_expected_missing" not in checks  # zero-rated rows ARE present

    # mandatory-tick gate: unticked confirmation blocks even with a warning note
    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation",
                    json={"warning_note": "Ledger verified with client"}, headers=H(ctx, "staff"))
    assert r.status_code == 409 and "exempt_apportionment" in r.json()["detail"]["reason"]
    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation",
                    json={"confirmations": ["exempt_apportionment", "zero_rated_evidence"],
                          "warning_note": "Ledger verified with client — profile to be updated"},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    tick = next(c for c in r.json()["computation"]["checks"] if c["id"] == "exempt_apportionment")
    assert tick["ticked_by_name"] == "Priya Nair" and tick["ticked_at"]
    assert any("Compliance confirmations ticked by Priya Nair" in e["text"]
               and "exempt_apportionment" in e["text"] and "zero_rated_evidence" in e["text"]
               for e in r.json()["events"])


def test_vat201_splits_completion_record_and_stars(client):
    ctx = setup_firm(client)
    cid, duty, f = setup_client_filing(client, ctx, ledger=CAT_LEDGER, register=CAT_REGISTER,
                                       profile_flags=ALL_YES)
    r = client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    c = r.json()["computation"]
    # profile-aware computation, VAT201-shaped
    assert c["profile_version"] == 1
    assert c["taxable_sales"] == 3000.0  # standard only
    assert c["per_emirate"] == {"Dubai": {"taxable_sales": 1000.0, "output_vat": 50.0, "rows": 1},
                                "Sharjah": {"taxable_sales": 2000.0, "output_vat": 100.0, "rows": 1}}
    assert c["zero_rated"] == {"sales": 5000.0, "rows": 1}
    assert c["exempt"] == {"sales": 3000.0, "rows": 1}
    assert c["margin"] == {"sales": 45000.0, "output_vat": 250.0, "rows": 1}
    assert c["rcm"] == {"output_vat": 100.0, "input_vat": 100.0, "rows": 1}
    assert c["output_vat"] == 500.0 and c["input_vat"] == 140.0
    assert c["net"] == 360.0 and c["position"] == "payable"
    # profile matches the data → confirmations only, no warnings
    kinds = {x["id"]: x["kind"] for x in c["checks"]}
    assert kinds == {"margin_confirmation": "confirmation", "rcm_confirmation": "confirmation",
                     "zero_rated_evidence": "confirmation", "exempt_apportionment": "confirmation"}

    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation", json={}, headers=H(ctx, "staff"))
    assert r.status_code == 409  # mandatory ticks missing
    r = client.post(f"/vat-engine/filings/{f['id']}/confirm-computation",
                    json={"confirmations": ["margin_confirmation", "rcm_confirmation",
                                            "zero_rated_evidence", "exempt_apportionment"]},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text

    # unchanged final flow: approval → FTA ack → duty completes → seal → stars
    r = client.post(f"/vat-engine/filings/{f['id']}/client-approval",
                    data={"basis": "email_approval", "note": "Approved by Mariam"}, headers=H(ctx, "staff"))
    assert r.status_code == 200
    r = client.post(f"/vat-engine/filings/{f['id']}/file-at-fta", data={"note": ""},
                    files=[("acknowledgement", ("ack.pdf", b"%PDF", "application/pdf"))],
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    duties = client.get("/duties", headers=H(ctx, "staff")).json()
    d = next(x for x in duties if x["id"] == duty["id"])
    rec = d["history"][-1]["record"]
    assert rec["zero-rated sales (AED)"] == "5,000.00"
    assert rec["exempt sales (AED)"] == "3,000.00"
    assert rec["net VAT (AED)"] == "360.00"
    assert d["history"][-1]["method"] == "proof"
    assert d["next_due"].startswith("2026-11-28")
    # sealed
    assert client.post(f"/vat-engine/filings/{f['id']}/draft-computation",
                       headers=H(ctx, "staff")).status_code == 409
    # stars flow through the existing performance pipeline
    emp = client.get("/performance/employees", headers=H(ctx, "admin")).json()["employees"]
    staff_row = next(e for e in emp if e["user_id"] == ctx["staff"]["id"])
    assert staff_row["duty_count"] >= 1
    assert any(ev["source"] == "duty" for ev in staff_row["recent_events"])


def test_stagger_drives_period_derivation_and_realigns(client):
    ctx = setup_firm(client)
    cid = make_client()
    duty = make_vat_duty(client, ctx, next_due="2026-08-28T00:00:00Z", client_id=cid)

    # profile recorded BEFORE opening: stagger Feb/May/Aug/Nov → period ends on the latest
    # stagger month-end before the Aug due month = 31 May
    r = client.post(f"/vat-engine/clients/{cid}/profile",
                    json={"business_category": "Trading", "tax_period_stagger": "feb_may_aug_nov",
                          "flags": {}}, headers=H(ctx, "staff"))
    assert r.status_code == 201, r.text
    f = open_filing(client, ctx, duty["id"])
    assert f["period_start"] == "2026-03-01" and f["period_end"] == "2026-05-31"
    assert f["prev_period_start"] == "2025-12-01"
    assert any("stagger Feb/May/Aug/Nov from profile v1" in e["text"] for e in f["events"])

    # stagger change re-aligns an open filing that hasn't started collecting
    r = client.patch(f"/vat-engine/clients/{cid}/profile",
                     json={"business_category": "Trading", "tax_period_stagger": "jan_apr_jul_oct",
                           "flags": {}}, headers=H(ctx, "manager"))
    assert r.status_code == 200, r.text
    f = client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()
    assert f["period_start"] == "2026-05-01" and f["period_end"] == "2026-07-31"
    assert any("re-aligned to the profile's tax period stagger (Jan/Apr/Jul/Oct)" in e["text"]
               for e in f["events"])
    # the version row records the stagger change with labels
    prof = client.get(f"/vat-engine/clients/{cid}/profile", headers=H(ctx, "staff")).json()
    row = next(c for c in prof["updated"][-1]["changes"] if c["field"] == "tax_period_stagger")
    assert row["old"] == "Feb/May/Aug/Nov" and row["new"] == "Jan/Apr/Jul/Oct"

    # monthly stagger → one-month periods regardless of the duty cadence
    cid2_duty = make_vat_duty(client, ctx, next_due="2026-08-28T00:00:00Z")
    f2 = open_filing(client, ctx, cid2_duty["id"])
    assert f2["period_start"] == "2026-05-01"  # no profile → cadence-based, unchanged behavior


def test_out_of_scope_category_flow(client):
    ctx = setup_firm(client)
    ledger = [
        ["S-1", date(2026, 5, 5), "Std Co", "T1", "Dubai", 1000, 50, "Output", "Standard (5%)"],
        ["DZ-1", date(2026, 5, 6), "JAFZA Co", "T2", "Dubai", 9000, 0, "Output",
         "Out of scope (designated zone)"],
    ]
    register = [
        ["S-1", date(2026, 5, 5), "Std Co", "Dubai", 1000, 50, "", ""],
        ["DZ-1", date(2026, 5, 6), "JAFZA Co", "Dubai", 9000, 0, "", "Out of scope (designated zone)"],
    ]
    # profile says designated_zone No → out-of-scope rows raise the mismatch warning
    _, _, f = setup_client_filing(client, ctx, ledger=ledger, register=register,
                                  profile_flags=flags(designated_zone="no"))
    assert {i["invoice_no"]: i["category"] for i in f["items"] if i["source"] == "ledger"} == \
        {"S-1": "standard", "DZ-1": "out_of_scope"}
    r = client.post(f"/vat-engine/filings/{f['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    c = r.json()["computation"]
    # out-of-scope reported outside the return boxes: no output VAT, no standard sales
    assert c["out_of_scope"] == {"sales": 9000.0, "rows": 1}
    assert c["taxable_sales"] == 1000.0 and c["output_vat"] == 50.0
    checks = {x["id"]: x["kind"] for x in c["checks"]}
    assert checks["out_of_scope_unexpected"] == "warning"

    # flip the profile: designated_zone = Not sure → treated as Yes, warning disappears
    r = client.patch(f"/vat-engine/clients/{f['client_id']}/profile",
                     json={"business_category": "Trading",
                           "flags": flags(designated_zone="not_sure")}, headers=H(ctx, "staff"))
    assert r.status_code == 200 and r.json()["flags"]["designated_zone"]["value"] == "not_sure"
    # re-draft (still at computation_draft → must go back? recon state is gone) — new filing instead:
    # simply assert the checks engine directly via a fresh draft on a second duty
    duty2 = make_vat_duty(client, ctx, client_id=f["client_id"])
    f2 = open_filing(client, ctx, duty2["id"])
    upload_ledger(client, ctx, f2["id"], fill_template(get_template(client, ctx, "ledger"), ledger))
    f2 = upload_register(client, ctx, f2["id"],
                         fill_template(get_template(client, ctx, "invoice-register"), register))
    r = client.post(f"/vat-engine/filings/{f2['id']}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    ids = [x["id"] for x in r.json()["computation"]["checks"]]
    assert "out_of_scope_unexpected" not in ids  # not_sure treated as Yes


def test_add_to_register_and_add_to_ledger_resolutions(client, monkeypatch):
    ctx = setup_firm(client)
    _, f = drive_to_reconciled(client, ctx)
    fid = f["id"]
    items = {(i["source"], i["invoice_no"]): i for i in f["items"]}
    assert f["unresolved_differences"] == 4  # l:INV-003, l:INV-004, i:INV-003, i:INV-900

    # (a) invoice found — evidence REQUIRED
    lid4 = items[("ledger", "INV-004")]["id"]
    form4 = {"invoice_no": "INV-004", "invoice_date": "2026-05-20", "party": "Delta Est",
             "emirate": "Ajman", "net": "500", "vat": "25", "note": "obtained from client by phone"}
    r = client.post(f"/vat-engine/filings/{fid}/items/{lid4}/add-to-register", data=form4,
                    headers=H(ctx, "staff"))
    assert r.status_code == 422 and "required as evidence" in r.json()["detail"]
    r = client.post(f"/vat-engine/filings/{fid}/items/{lid4}/add-to-register", data=form4,
                    files=[("evidence", ("inv-004.pdf", b"%PDF obtained", "application/pdf"))],
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    f = r.json()
    by = {(i["source"], i["invoice_no"], i["origin"]): i for i in f["items"]}
    assert by[("ledger", "INV-004", "register")]["bucket"] == "matched"
    added = by[("invoice", "INV-004", "added_at_recon")]
    assert added["bucket"] == "matched" and "evidence: inv-004.pdf" in added["notes"]
    assert f["unresolved_differences"] == 3  # gate re-evaluated immediately
    assert any("Difference resolved — invoice INV-004 obtained and added to register by Priya Nair, "
               "evidence attached" in e["text"] for e in f["events"])

    # gate REOPENS when the added invoice doesn't actually match
    lid3 = items[("ledger", "INV-003")]["id"]
    r = client.post(f"/vat-engine/filings/{fid}/items/{lid3}/add-to-register",
                    data={**form4, "invoice_no": "INV-003", "vat": "999"},
                    files=[("evidence", ("inv-003.pdf", b"%PDF", "application/pdf"))],
                    headers=H(ctx, "staff"))
    assert r.status_code == 200
    f = r.json()
    assert f["unresolved_differences"] == 4  # new invoice_only difference opened
    assert any("does NOT match the ledger row" in e["text"] for e in f["events"])

    # (b) missing from client ledger — correction note MANDATORY (schema-enforced)
    iid900 = items[("invoice", "INV-900")]["id"]
    body900 = {"invoice_no": "INV-900", "invoice_date": "2026-06-20", "party": "Mystery Co",
               "emirate": "Dubai", "net": 700, "vat": 35, "type": "Output", "category": "standard",
               "note": ""}
    r = client.post(f"/vat-engine/filings/{fid}/items/{iid900}/add-to-ledger", json=body900,
                    headers=H(ctx, "staff"))
    assert r.status_code == 422  # empty note refused
    body900["note"] = "client ledger omitted this invoice — to be booked in client's records"
    r = client.post(f"/vat-engine/filings/{fid}/items/{iid900}/add-to-ledger", json=body900,
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    f = r.json()
    corr = next(i for i in f["items"] if i["origin"] == "ledger_correction")
    assert corr["source"] == "ledger" and corr["bucket"] == "matched" and corr["type"] == "Output"
    assert f["unresolved_differences"] == 3
    assert any("missing ledger entry added by Priya Nair for invoice INV-900" in e["text"]
               for e in f["events"])

    # close the remaining differences → gate opens
    remaining = [i for i in f["items"] if i["bucket"] in ("ledger_only", "invoice_only")
                 and (i.get("resolution") or {}).get("action") not in ("excluded", "resolved")]
    assert len(remaining) == 3
    for i in remaining:
        r = client.post(f"/vat-engine/filings/{fid}/items/{i['id']}/exclude",
                        json={"reason": f"not part of this filing — {i['invoice_no']}"},
                        headers=H(ctx, "staff"))
        assert r.status_code == 200
    r = client.post(f"/vat-engine/filings/{fid}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text

    # the regenerated workbook reflects resolutions + the Ledger corrections sheet
    from openpyxl import load_workbook
    r = client.get(f"/vat-engine/filings/{fid}/recon-workbook", headers=H(ctx, "staff"))
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert "Ledger corrections" in wb.sheetnames
    cw = wb["Ledger corrections"]
    corr_rows = [[cw.cell(row=r_, column=c).value for c in range(1, 14)]
                 for r_ in range(3, cw.max_row + 1)]
    assert any(row[3] == "INV-900" and "to be booked in client's records" in (row[12] or "")
               for row in corr_rows)
    diff_ws = wb["Differences"]
    assert diff_ws.cell(row=1, column=12).value == "Resolution"
    origins = {wb["Matched"].cell(row=r_, column=2).value for r_ in range(2, wb["Matched"].max_row + 1)}
    assert {"Added at recon", "Ledger correction"} <= origins

    # the computation email tells the client to book the correction
    sent = {}
    monkeypatch.setattr(vat_engine.emails, "send_client",
                        lambda to, subject, body: sent.update(to=to, body=body))
    r = client.post(f"/vat-engine/filings/{fid}/confirm-computation", json={}, headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    r = client.post(f"/vat-engine/filings/{fid}/send-computation",
                    json={"to": "accounts@gulfhorizon.ae", "subject": "VAT computation",
                          "body": "Please approve."}, headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    assert "The following invoices were added to the VAT workings and should be booked in your ledger" in sent["body"]
    assert "INV-900" in sent["body"] and "to be booked in client's records" in sent["body"]


def test_working_paper_sheets_and_adjustment_loop(client):
    from openpyxl import load_workbook

    ctx = setup_firm(client)
    cid, duty, f = setup_client_filing(client, ctx, ledger=CAT_LEDGER, register=CAT_REGISTER,
                                       profile_flags=ALL_YES)
    fid = f["id"]
    r = client.post(f"/vat-engine/filings/{fid}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text

    def get_workbook():
        resp = client.get(f"/vat-engine/filings/{fid}/recon-workbook", headers=H(ctx, "staff"))
        assert resp.status_code == 200
        return load_workbook(io.BytesIO(resp.content))

    def rows_of(ws):
        return [[c.value for c in row] for row in ws.iter_rows()]

    # 1 — all sheets present, computation sheets after the recon sheets
    wb = get_workbook()
    assert wb.sheetnames == ["Summary", "Matched", "Differences", "Excluded", "Ledger corrections",
                             "VAT Computation", "Computation Detail"]
    comp = wb["VAT Computation"]
    rows = rows_of(comp)
    assert rows[1][0] == "This workbook is generated from Baton; corrections are made in Baton."
    assert "Gulf Horizon Trading LLC" in rows[2][0] and "CL-001" in rows[2][0]
    assert rows[4][0] == "Client VAT profile applied: v1"
    labels = [r_[0] for r_ in rows]
    dubai = rows[labels.index("Standard-rated — Dubai")]
    assert dubai[1] == 1000.0 and dubai[2] == 50.0 and dubai[3] == 1
    sub = rows[labels.index("Standard-rated subtotal")]
    assert str(sub[1]).startswith("=SUM(") and str(sub[2]).startswith("=SUM(")
    assert "Box 1" in sub[4]
    zr = rows[labels.index("Zero-rated supplies")]
    assert zr[1] == 5000.0 and "no output VAT (Art. 45)" in zr[4]
    out_total = rows[labels.index("Output VAT (total)")]
    assert str(out_total[2]).startswith("=C")
    inp = rows[labels.index("Input VAT (recoverable)")]
    assert inp[2] == 140.0 and "Art. 55" in inp[4]
    net_row = rows[labels.index("NET VAT PAYABLE")]
    assert str(net_row[2]).startswith("=C")
    assert comp.freeze_panes == "A8"
    # unticked footer — the confirmations are pending
    assert any(r_[4] and "NOT YET TICKED" in str(r_[4]) for r_ in rows)

    detail = wb["Computation Detail"]
    drows = rows_of(detail)
    assert drows[1][:6] == ["Invoice No", "Date", "Party", "Emirate", "Category", "Type"]
    data = [r_ for r_ in drows[2:] if r_[0] and r_[0] != "TOTAL"]
    assert len(data) == 7 and all(r_[8] == "Ledger" for r_ in data)
    total = next(r_ for r_ in drows if r_[0] == "TOTAL")
    assert str(total[6]).startswith("=SUM(") and str(total[7]).startswith("=SUM(")
    assert detail.freeze_panes == "A3"

    # 2 — adjustment: edit re-drafts live, resets ticks, logs
    items = {i["invoice_no"]: i for i in r.json()["items"] if i["source"] == "ledger"}
    r = client.post(f"/vat-engine/filings/{fid}/items/{items['S-1']['id']}/adjust",
                    json={"action": "edit", "emirate": "Sharjah",
                          "reason": "invoice shows place of supply Sharjah"},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    f2 = r.json()
    c2 = f2["computation"]
    assert c2["per_emirate"] == {"Sharjah": {"taxable_sales": 3000.0, "output_vat": 150.0, "rows": 2}}
    assert all(not x.get("ticked_by") for x in c2["checks"])  # confirmations RESET
    assert any('Computation adjustment by Priya Nair: S-1 emirate Dubai→Sharjah — reason: '
               '"invoice shows place of supply Sharjah"' in e["text"] for e in f2["events"])
    # confirm without re-ticking → blocked (numbers changed)
    r = client.post(f"/vat-engine/filings/{fid}/confirm-computation", json={}, headers=H(ctx, "staff"))
    assert r.status_code == 409

    # exclude an input row → input VAT drops to the RCM-only 100
    r = client.post(f"/vat-engine/filings/{fid}/items/{items['P-1']['id']}/adjust",
                    json={"action": "exclude", "reason": "duplicate of a Q1 booking"},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    assert r.json()["computation"]["input_vat"] == 100.0
    assert r.json()["computation"]["counts"]["included"] == 6

    # jump back to Stage 3 and re-draft — the exclusion persists
    r = client.post(f"/vat-engine/filings/{fid}/reopen-reconciliation",
                    json={"reason": "double-check a difference"}, headers=H(ctx, "staff"))
    assert r.status_code == 200 and r.json()["status"] == "reconciled" and r.json()["computation"] is None
    r = client.post(f"/vat-engine/filings/{fid}/draft-computation", headers=H(ctx, "staff"))
    assert r.status_code == 200 and r.json()["computation"]["input_vat"] == 100.0

    # 3 — confirm with ticks → the footer names the confirmer; Detail notes the adjustment
    ticks = [x["id"] for x in r.json()["computation"]["checks"] if x["kind"] == "confirmation"]
    r = client.post(f"/vat-engine/filings/{fid}/confirm-computation",
                    json={"confirmations": ticks}, headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    wb = get_workbook()
    rows = rows_of(wb["VAT Computation"])
    assert any(r_[4] and "ticked by Priya Nair" in str(r_[4]) for r_ in rows)
    drows = rows_of(wb["Computation Detail"])
    s1 = next(r_ for r_ in drows if r_[0] == "S-1")
    assert "adjusted by Priya Nair: emirate Dubai→Sharjah — invoice shows place of supply Sharjah" in s1[9]


# ---------- AI invoice extraction (mandatory human review) ----------

from app.routers import vat_engine  # noqa: E402


def fake_fields(no="AI-100", date_="2026-05-12", party="Acme LLC", emirate="Dubai",
                net=1000, vat=50, curr="AED", conf="high"):
    mk = lambda v: {"value": v, "confidence": conf}  # noqa: E731
    return {"invoice_no": mk(no), "invoice_date": mk(date_), "party": mk(party),
            "emirate": mk(emirate), "net_amount": mk(net), "vat_amount": mk(vat),
            "currency": mk(curr)}


def post_extract(client, ctx, fid, files, expect=200):
    r = client.post(f"/vat-engine/filings/{fid}/invoices/extract",
                    files=[("files", (n, data, "application/pdf")) for n, data in files],
                    headers=H(ctx, "staff"))
    assert r.status_code == expect, r.text
    return r.json()


def extraction_filing(client, ctx, ledger_rows):
    cid = make_client()
    duty = make_vat_duty(client, ctx, client_id=cid)
    client.post(f"/vat-engine/clients/{cid}/profile",
                json={"business_category": "Trading", "flags": {}}, headers=H(ctx, "staff"))
    f = open_filing(client, ctx, duty["id"])
    upload_ledger(client, ctx, f["id"], fill_template(get_template(client, ctx, "ledger"), ledger_rows))
    return client.get(f"/vat-engine/filings/{f['id']}", headers=H(ctx, "staff")).json()


def test_ai_extraction_review_gate_and_mixed_recon(client, monkeypatch):
    ctx = setup_firm(client)
    ledger = [
        ["AI-100", date(2026, 5, 12), "Acme LLC", "T1", "Dubai", 1000, 50, "Output", ""],
        ["REG-1", date(2026, 5, 15), "Reg Co", "T2", "Dubai", 500, 25, "Output", ""],
    ]
    f = extraction_filing(client, ctx, ledger)

    def fake(data, media_type):
        if b"nullish" in data:
            out = fake_fields(no=None, net=None, conf="low")
            return out
        return fake_fields()
    monkeypatch.setattr(vat_engine, "_extract_invoice", fake)

    out = post_extract(client, ctx, f["id"], [("inv-ai-100.pdf", b"%PDF good"),
                                              ("blurry.pdf", b"%PDF nullish")])
    assert [x["status"] for x in out["results"]] == ["extracted", "extracted"]
    drafts = out["filing"]["extraction_drafts"]
    assert len(drafts) == 2
    nullish = next(d for d in drafts if d["file_name"] == "blurry.pdf")
    assert nullish["fields"]["invoice_no"]["value"] is None
    assert nullish["fields"]["net_amount"]["confidence"] == "low"
    # the source documents are kept as evidence files
    assert all(d["file_id"] for d in drafts)

    # REVIEW GATE: drafts are not items — the register upload reconciles without them
    register = [["REG-1", date(2026, 5, 15), "Reg Co", "Dubai", 500, 25, "", ""]]
    f2 = upload_register(client, ctx, f["id"],
                         fill_template(get_template(client, ctx, "invoice-register"), register))
    assert {i["invoice_no"] for i in f2["items"] if i["source"] == "invoice"} == {"REG-1"}
    by = {(i["source"], i["invoice_no"]): i for i in f2["items"]}
    assert by[("ledger", "AI-100")]["bucket"] == "ledger_only"  # AI draft not reconciled yet

    # confirm the good draft with one correction (party) → joins recon as origin=ai_extracted
    good = next(d for d in drafts if d["file_name"] == "inv-ai-100.pdf")
    r = client.post(f"/vat-engine/filings/{f['id']}/invoices/confirm-extracted",
                    json={"rows": [{"draft_id": good["id"], "invoice_no": "AI-100",
                                    "invoice_date": "2026-05-12", "party": "Acme LLC (Dubai branch)",
                                    "emirate": "Dubai", "net": 1000, "vat": 50, "currency": "AED"}]},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    f3 = r.json()
    item = next(i for i in f3["items"] if i["invoice_no"] == "AI-100" and i["source"] == "invoice")
    assert item["origin"] == "ai_extracted" and item["bucket"] == "matched"
    assert "AI-extracted from inv-ai-100.pdf" in item["notes"]
    # mixed recon: register + AI-extracted both matched
    by2 = {(i["source"], i["invoice_no"]): i["bucket"] for i in f3["items"]}
    assert by2[("invoice", "REG-1")] == "matched" and by2[("ledger", "AI-100")] == "matched"
    assert any("1 invoice(s) extracted by AI, reviewed and confirmed by Priya Nair — 1 field(s) corrected"
               in e["text"] for e in f3["events"])
    # the unconfirmed draft stays a draft — never reconciles
    assert next(d for d in f3["extraction_drafts"] if d["file_name"] == "blurry.pdf")["status"] == "extracted"
    # the recon workbook rows carry the origin
    from openpyxl import load_workbook
    from app import blobs
    with engine.begin() as conn:
        blob_path = conn.execute(sql("SELECT blob_path FROM files WHERE id = :i"),
                                 {"i": f3["recon"]["excel_file_id"]}).scalar()
    wb = load_workbook(io.BytesIO(blobs.read_blob(blob_path)))
    ws = wb["Matched"]
    assert ws.cell(row=1, column=2).value == "Origin"
    origins = {ws.cell(row=r_, column=2).value for r_ in range(2, ws.max_row + 1)}
    assert {"Ledger", "AI-extracted", "Register"} <= origins


def test_ai_extraction_batch_failure_isolation(client, monkeypatch):
    ctx = setup_firm(client)
    f = extraction_filing(client, ctx, [["X-1", date(2026, 5, 5), "P", "T", "Dubai", 100, 5, "Output", ""]])

    def fake(data, media_type):
        if b"corrupt" in data:
            raise RuntimeError("model choked")
        return fake_fields()
    monkeypatch.setattr(vat_engine, "_extract_invoice", fake)
    out = post_extract(client, ctx, f["id"], [("ok.pdf", b"%PDF fine"), ("bad.pdf", b"%PDF corrupt")])
    statuses = {x["file_name"]: x for x in out["results"]}
    assert statuses["ok.pdf"]["status"] == "extracted"
    assert statuses["bad.pdf"]["status"] == "failed"
    assert "extraction failed — enter manually" in statuses["bad.pdf"]["error"]
    # unsupported type fails without touching the API
    out = post_extract(client, ctx, f["id"], [("notes.txt", b"hello")])
    assert out["results"][0]["status"] == "failed"


def test_ai_extraction_max_files_guard(client, monkeypatch):
    ctx = setup_firm(client)
    f = extraction_filing(client, ctx, [["X-1", date(2026, 5, 5), "P", "T", "Dubai", 100, 5, "Output", ""]])
    monkeypatch.setenv("VAT_EXTRACT_MAX_FILES", "2")
    r = client.post(f"/vat-engine/filings/{f['id']}/invoices/extract",
                    files=[("files", (f"i{i}.pdf", b"%PDF", "application/pdf")) for i in range(3)],
                    headers=H(ctx, "staff"))
    assert r.status_code == 422
    assert "limit of 2" in r.json()["detail"]


def test_ai_extraction_non_aed_requires_conversion_note(client, monkeypatch):
    ctx = setup_firm(client)
    f = extraction_filing(client, ctx, [["USD-1", date(2026, 5, 5), "P", "T", "Dubai", 3673, 183.65, "Output", ""]])
    monkeypatch.setattr(vat_engine, "_extract_invoice",
                        lambda data, mt: fake_fields(no="USD-1", net=1000, vat=50, curr="USD"))
    out = post_extract(client, ctx, f["id"], [("usd.pdf", b"%PDF")])
    draft_id = out["results"][0]["draft_id"]
    row = {"draft_id": draft_id, "invoice_no": "USD-1", "invoice_date": "2026-05-05",
           "party": "P", "emirate": "Dubai", "net": 3673.0, "vat": 183.65, "currency": "USD"}
    r = client.post(f"/vat-engine/filings/{f['id']}/invoices/confirm-extracted",
                    json={"rows": [row]}, headers=H(ctx, "staff"))
    assert r.status_code == 409 and "manual-conversion note is mandatory" in r.json()["detail"]["reason"]
    r = client.post(f"/vat-engine/filings/{f['id']}/invoices/confirm-extracted",
                    json={"rows": [{**row, "conversion_note": "CB rate 3.6730 on invoice date"}]},
                    headers=H(ctx, "staff"))
    assert r.status_code == 200, r.text
    item = next(i for i in r.json()["items"] if i["invoice_no"] == "USD-1" and i["source"] == "invoice")
    assert "USD converted manually — CB rate 3.6730" in item["notes"]


def test_tenancy_isolation(client):
    ctx = setup_firm(client)
    duty = make_vat_duty(client, ctx)
    f = open_filing(client, ctx, duty["id"])

    boot_b = bootstrap_tenant(client, email="hello@betabooks.ae", user_email_domain="betabooks.ae")
    admin_b = next(u for u in boot_b["users"] if u["role"] == "Admin")
    tokens_b = login_after_reset(client, admin_b["email"], admin_b["temp_password"])
    headers_b = {"Authorization": f"Bearer {tokens_b['access_token']}"}
    assert client.get(f"/vat-engine/filings/{f['id']}", headers=headers_b).status_code == 404
    assert f["id"] not in {x["id"] for x in client.get("/vat-engine/filings", headers=headers_b).json()}
